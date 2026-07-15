"""Create a per-category comparison matrix from the reference assignment workbook.

Default input:
    reports/CFC_Reference_Folder_Assignment.xlsx

Default output:
    reports/CFC_Category_Comparison_Matrix.xlsx
"""

from __future__ import annotations

import argparse
import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from cfc_research_library import CATEGORY_COLORS, SECTIONS


DEFAULT_INPUT = "reports/CFC_Reference_Folder_Assignment.xlsx"
DEFAULT_OUTPUT = "reports/CFC_Category_Comparison_Matrix.xlsx"
DEFAULT_SHEET = "Reference_Assignments"

CATEGORIES = [
    category
    for category in SECTIONS
    if category not in {"Historical Articles", "Conferences"}
]

CATEGORY_ALIASES = {
    "allergy": "Allergy and Immunology",
    "allergy and immunology": "Allergy and Immunology",
    "cardio": "Cardiology",
    "cardiology": "Cardiology",
    "derm": "Dermatology",
    "dermatology": "Dermatology",
    "development": "Development and Behavior",
    "development and behavior": "Development and Behavior",
    "behavior": "Development and Behavior",
    "endocrinology": "Endocrinology",
    "gastro": "Gastroenterology",
    "gastroenterology": "Gastroenterology",
    "general": "General and Reviews",
    "general and reviews": "General and Reviews",
    "reviews": "General and Reviews",
    "genetic": "Genetics",
    "genetics": "Genetics",
    "growth": "Growth",
    "gynecology": "Gynecology",
    "neuro": "Neurology",
    "neuology": "Neurology",
    "neurology": "Neurology",
    "oncology": "Oncology",
    "ophthalmology": "Ophthalmology",
    "orthopedic": "Orthopedic",
    "orthopedics": "Orthopedic",
    "otolaryngology": "Otolaryngology",
    "pulmonology": "Pulmonology",
    "research": "Research Studies",
    "research studies": "Research Studies",
    "seizure": "Seizures",
    "seizures": "Seizures",
    "treatment": "Treatments",
    "treatments": "Treatments",
    "exclude": "Excluded",
    "excluded": "Excluded",
}


def split_categories(value: object) -> set[str]:
    if value is None or pd.isna(value):
        return set()
    text = str(value).strip().lower()
    if not text:
        return set()
    protected = {
        "allergy and immunology": "allergy_immunology",
        "development and behavior": "development_behavior",
        "general and reviews": "general_reviews",
    }
    for phrase, token in protected.items():
        text = text.replace(phrase, token)
    text = re.sub(r"\badditional\s*:\s*", ";", text)
    text = text.replace("&", ";")
    text = re.sub(r"\band\b", ";", text)
    text = text.replace(",", ";").replace("/", ";").replace("|", ";")
    for phrase, token in protected.items():
        text = text.replace(token, phrase)
    parts = [re.sub(r"\s+", " ", part).strip(" .;:?") for part in text.split(";")]
    categories: set[str] = set()
    for part in parts:
        if not part:
            continue
        if part in CATEGORY_ALIASES:
            categories.add(CATEGORY_ALIASES[part])
            continue
        for alias, category in CATEGORY_ALIASES.items():
            if re.search(rf"\b{re.escape(alias)}\b", part):
                categories.add(category)
    return {category for category in categories if category in CATEGORIES or category == "Excluded"}


def find_column(df: pd.DataFrame, candidates: tuple[str, ...]) -> str | None:
    normalized = {re.sub(r"[^a-z0-9]+", " ", str(column).lower()).strip(): column for column in df.columns}
    for candidate in candidates:
        key = re.sub(r"[^a-z0-9]+", " ", candidate.lower()).strip()
        if key in normalized:
            return normalized[key]
    return None


def normalize_decision(value: object) -> str:
    text = re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()
    if not text:
        return ""
    if any(term in text for term in ("not relevant", "not approved", "screen out", "excluded", "exclude")):
        return "Excluded"
    if any(term in text for term in ("relevant", "approved", "screen in", "included", "include")):
        return "Included"
    if "review" in text or "possible" in text:
        return "Needs review"
    return str(value or "").strip()


def category_text(categories: set[str]) -> str:
    return "; ".join(category for category in CATEGORIES if category in categories)


def build_article_details(df: pd.DataFrame) -> pd.DataFrame:
    existing_col = find_column(df, ("Primary_Category", "Primary Category", "Existing_Category", "Existing Category"))
    matched_col = find_column(df, ("Matched_Categories", "Matched Categories", "Existing_Matched_Categories"))
    api_col = find_column(df, ("API_Suggested_Category", "API Suggested Category", "OpenAI Assigned Category"))
    api_secondary_col = find_column(df, ("API_Secondary_Category", "OpenAI Secondary Categories", "Secondary Categories"))
    decision_col = find_column(df, ("API_Relevance_Decision", "OpenAI_Relevance_Decision", "Relevance"))
    pmid_col = find_column(df, ("PMID", "PubMed ID", "PubMed_ID"))
    title_col = find_column(df, ("Title", "Article Title"))
    analysis_col = find_column(df, ("API_Analysis", "OpenAI_Rationale", "Rationale"))
    confidence_col = find_column(df, ("API_Confidence", "OpenAI_Confidence", "Confidence"))
    url_col = find_column(df, ("PubMed_URL", "PubMed URL", "URL"))

    rows = []
    for _, row in df.iterrows():
        existing_raw = "; ".join(
            str(row.get(column, "") or "").strip()
            for column in (existing_col, matched_col)
            if column and str(row.get(column, "") or "").strip()
        )
        api_raw = "; ".join(
            str(row.get(column, "") or "").strip()
            for column in (api_col, api_secondary_col)
            if column and str(row.get(column, "") or "").strip()
        )
        existing_categories = split_categories(existing_raw)
        api_categories = split_categories(api_raw)
        decision = normalize_decision(row.get(decision_col, "")) if decision_col else ""
        if decision == "Excluded" and not api_categories:
            api_categories = {"Excluded"}

        rows.append(
            {
                "PMID": row.get(pmid_col, "") if pmid_col else "",
                "Title": row.get(title_col, "") if title_col else "",
                "Existing_Category": row.get(existing_col, "") if existing_col else "",
                "Existing_Matched_Categories": row.get(matched_col, "") if matched_col else "",
                "Existing_Categories_Normalized": category_text(existing_categories),
                "API_Suggested_Category": row.get(api_col, "") if api_col else "",
                "API_Secondary_Category": row.get(api_secondary_col, "") if api_secondary_col else "",
                "API_Categories_Normalized": category_text(api_categories),
                "API_Relevance_Decision": decision,
                "Category_Exact_Match": "Yes" if existing_categories and api_categories and existing_categories == api_categories else ("No" if existing_categories and api_categories else ""),
                "Any_Category_Overlap": "Yes" if existing_categories and api_categories and bool(existing_categories & api_categories) else ("No" if existing_categories and api_categories else ""),
                "AI_Subset_of_Existing": "Yes" if existing_categories and api_categories and api_categories.issubset(existing_categories) else ("No" if existing_categories and api_categories else ""),
                "Existing_Subset_of_AI": "Yes" if existing_categories and api_categories and existing_categories.issubset(api_categories) else ("No" if existing_categories and api_categories else ""),
                "API_Confidence": row.get(confidence_col, "") if confidence_col else "",
                "API_Analysis": row.get(analysis_col, "") if analysis_col else "",
                "PubMed_URL": row.get(url_col, "") if url_col else "",
                "_existing_set": existing_categories,
                "_api_set": api_categories,
            }
        )
    return pd.DataFrame(rows)


def build_category_2x2(details: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for category in CATEGORIES:
        tp = fp = fn = tn = 0
        for _, row in details.iterrows():
            human_yes = category in row["_existing_set"]
            api_yes = category in row["_api_set"]
            if human_yes and api_yes:
                tp += 1
            elif not human_yes and api_yes:
                fp += 1
            elif human_yes and not api_yes:
                fn += 1
            else:
                tn += 1
        total = tp + fp + fn + tn
        precision = tp / (tp + fp) if tp + fp else ""
        recall = tp / (tp + fn) if tp + fn else ""
        f1 = 2 * precision * recall / (precision + recall) if precision != "" and recall != "" and precision + recall else ""
        rows.append(
            {
                "Category": category,
                "True_Positive": tp,
                "False_Positive": fp,
                "False_Negative": fn,
                "True_Negative": tn,
                "Total": total,
                "Agreement_Rate": (tp + tn) / total if total else "",
                "Precision": precision,
                "Recall": recall,
                "F1_Score": f1,
                "Specificity": tn / (tn + fp) if tn + fp else "",
                "Needs_Review": fp + fn,
            }
        )
    return pd.DataFrame(rows)


def build_overall_metrics(details: pd.DataFrame) -> pd.DataFrame:
    counted = details[
        details["Existing_Categories_Normalized"].astype(str).str.len().gt(0)
        & details["API_Categories_Normalized"].astype(str).str.len().gt(0)
    ].copy()
    total = len(counted)
    if not total:
        return pd.DataFrame()
    return pd.DataFrame(
        [
            {"Metric": "Rows counted", "Value": total},
            {"Metric": "Exact category-list match", "Value": int((counted["Category_Exact_Match"] == "Yes").sum()), "Rate": (counted["Category_Exact_Match"] == "Yes").mean()},
            {"Metric": "Any category overlap", "Value": int((counted["Any_Category_Overlap"] == "Yes").sum()), "Rate": (counted["Any_Category_Overlap"] == "Yes").mean()},
            {"Metric": "AI category list subset of existing list", "Value": int((counted["AI_Subset_of_Existing"] == "Yes").sum()), "Rate": (counted["AI_Subset_of_Existing"] == "Yes").mean()},
            {"Metric": "Existing category list subset of AI list", "Value": int((counted["Existing_Subset_of_AI"] == "Yes").sum()), "Rate": (counted["Existing_Subset_of_AI"] == "Yes").mean()},
        ]
    )


def criteria_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Category": section.name,
                "Description": section.description,
                "Inclusion": section.inclusion,
                "Exclusion": section.exclusion,
            }
            for name, section in SECTIONS.items()
            if name in CATEGORIES
        ]
    )


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    match_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    mismatch_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
        headers = [cell.value for cell in ws[1]]
        max_row = max(ws.max_row, 2)
        for header in ("Category_Exact_Match", "Any_Category_Overlap", "AI_Subset_of_Existing", "Existing_Subset_of_AI"):
            if header in headers:
                letter = get_column_letter(headers.index(header) + 1)
                ws.conditional_formatting.add(f"{letter}2:{letter}{max_row}", FormulaRule(formula=[f'${letter}2="Yes"'], fill=match_fill))
                ws.conditional_formatting.add(f"{letter}2:{letter}{max_row}", FormulaRule(formula=[f'${letter}2="No"'], fill=mismatch_fill))
        if "Category" in headers:
            col = headers.index("Category") + 1
            for row in range(2, max_row + 1):
                category = str(ws.cell(row=row, column=col).value or "")
                color = CATEGORY_COLORS.get(category, "FFFFFF")
                ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        widths = {
            "Title": 65,
            "Existing_Categories_Normalized": 34,
            "API_Categories_Normalized": 34,
            "API_Analysis": 70,
            "PubMed_URL": 45,
            "Inclusion": 70,
            "Exclusion": 70,
            "Description": 55,
        }
        for idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(str(header), 22)
    wb.save(path)


def write_workbook(output_path: Path, details: pd.DataFrame, input_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    visible_details = details.drop(columns=["_existing_set", "_api_set"])
    category_summary = build_category_2x2(details)
    disagreements = visible_details[
        (visible_details["Any_Category_Overlap"] == "No")
        | (visible_details["Category_Exact_Match"] == "No")
    ].copy()
    run_log = pd.DataFrame(
        [
            {"Field": "Input workbook", "Value": str(input_path)},
            {"Field": "Input sheet", "Value": DEFAULT_SHEET},
            {"Field": "Output workbook", "Value": str(output_path)},
            {"Field": "Generated", "Value": time.ctime()},
            {"Field": "Category normalization", "Value": "Uses human-aligned category aliases; splits semicolon/comma/slash/and lists; excludes Historical Articles and Conferences."},
            {"Field": "Criteria alignment", "Value": "Aligned with stricter CFC rules: meaningful CFC discussion required; passing mentions/figure-only/reference-only CFC allusions should be excluded upstream."},
            {"Field": "OpenAI model", "Value": "No OpenAI call is made by this matrix script."},
        ]
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        build_overall_metrics(details).to_excel(writer, sheet_name="Overall_Metrics", index=False)
        category_summary.to_excel(writer, sheet_name="2x2_By_Category", index=False)
        visible_details.to_excel(writer, sheet_name="Article_Details", index=False)
        disagreements.to_excel(writer, sheet_name="Disagreements", index=False)
        criteria_frame().to_excel(writer, sheet_name="Criteria", index=False)
        run_log.to_excel(writer, sheet_name="Run_Log", index=False)
    format_workbook(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a category comparison matrix workbook.")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Reference assignment workbook to analyze.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Category comparison workbook to create.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Sheet in the reference workbook to read.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    df = pd.read_excel(input_path, sheet_name=args.sheet, dtype=str)
    details = build_article_details(df)
    write_workbook(output_path, details, input_path)
    print(f"Category comparison workbook written: {output_path.resolve()}")


if __name__ == "__main__":
    main()
