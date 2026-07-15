"""Update a CFC syndrome literature review workbook from PubMed and Zotero.

This replaces the original Colab-only notebook with a reusable local program:

    python cfc_research_library.py --category Dermatology --output reports/Dermatology_Master_Review_Report.xlsx

Credentials are read from environment variables so API keys are not stored in code.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
import textwrap
import time
from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable, TypeAlias
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlparse
from urllib.request import Request, urlopen, urlretrieve

pd = None
Entrez = None
zotero = None
SentenceTransformer = None
util = None

DataFrame: TypeAlias = Any


CFC_TERMS = (
    "cardiofaciocutaneous",
    "cardio-facio-cutaneous",
    "cardio-facio-cutaneous syndrome",
    "cardiofaciocutaneous syndrome",
    "CFC syndrome",
)

CFC_GENES = ("BRAF", "MAP2K1", "MAP2K2", "KRAS")

REVIEWER_DECISIONS = ("Needs reviewed", "Approved", "Not approved")

DEFAULT_SCREENING_HISTORY_URL = (
    "https://docs.google.com/spreadsheets/d/1BUvWcV6XgYiOL3cCrYAHjkccb24C38OK/"
    "edit?usp=sharing&ouid=106518116377917721454&rtpof=true&sd=true"
)

CATEGORY_COLORS = {
    "Allergy and Immunology": "D9EAD3",
    "Cardiology": "F4CCCC",
    "Dermatology": "FCE5CD",
    "Development and Behavior": "D9D2E9",
    "Endocrinology": "D0E0E3",
    "Gastroenterology": "FFF2CC",
    "General and Reviews": "EADCF8",
    "Genetics": "CFE2F3",
    "Growth": "E2F0CB",
    "Gynecology": "FCE4EC",
    "Neurology": "C9DAF8",
    "Oncology": "E6B8AF",
    "Ophthalmology": "D9EAD3",
    "Orthopedic": "EAD7C2",
    "Otolaryngology": "D0E0E3",
    "Pulmonology": "D9EAF7",
    "Research Studies": "E6E6E6",
    "Seizures": "D9D2E9",
    "Treatments": "D5A6BD",
    "Historical Articles": "EFEFEF",
    "Conferences": "D9D9D9",
}

OTHER_RASOPATHY_SIGNALS = (
    "costello syndrome",
    "hras-positive",
    "hras positive",
    "noonan syndrome",
    "legius syndrome",
    "neurofibromatosis type 1",
)

STRONG_CFC_SIGNALS = (
    "cardiofaciocutaneous syndrome",
    "cardio-facio-cutaneous syndrome",
    "cfc syndrome",
    "cardiofaciocutaneous",
    "cardio-facio-cutaneous",
)

GENERAL_INCLUSION_CRITERIA = [
    "Directly discusses cardiofaciocutaneous syndrome, CFC syndrome, or a CFC-specific subgroup within RASopathy research.",
    "Includes confirmed or strongly suspected CFC cases, or uses CFC-associated RAS/MAPK variants to model relevant clinical biology.",
    "Provides clinical, genetic, mechanistic, management, or review evidence relevant to the selected Zotero library section.",
]

GENERAL_EXCLUSION_CRITERIA = [
    "Focuses only on Noonan, Costello, or other RASopathies without CFC-specific data or interpretation.",
    "Mentions RAS/MAPK biology without a clear CFC clinical, genetic, or phenotypic connection.",
    "Is unrelated to humans, CFC-relevant model systems, clinical care, diagnosis, treatment, or literature synthesis.",
]

EXCLUDED_COMPARISON_FOLDERS = {"Exclusion", "Excluded", "Historical Articles", "Conferences"}
MAX_ZOTERO_EXAMPLES_PER_CATEGORY = 5

CATEGORY_ALIASES = {
    "allergy": "Allergy and Immunology",
    "allergy and immunology": "Allergy and Immunology",
    "cardio": "Cardiology",
    "cardiology": "Cardiology",
    "derm": "Dermatology",
    "dermatology": "Dermatology",
    "development": "Development and Behavior",
    "development and behavior": "Development and Behavior",
    "behavior": "Development and Behavior",
    "endocrinology": "Endocrinology",
    "gastro": "Gastroenterology",
    "gastroenterology": "Gastroenterology",
    "general": "General and Reviews",
    "general_and_reviews": "General and Reviews",
    "general and reviews": "General and Reviews",
    "reviews": "General and Reviews",
    "genetic": "Genetics",
    "genetics": "Genetics",
    "growth": "Growth",
    "gynecology": "Gynecology",
    "neuro": "Neurology",
    "neuology": "Neurology",
    "neurology": "Neurology",
    "oncology": "Oncology",
    "ophthalmology": "Ophthalmology",
    "orthopedic": "Orthopedic",
    "orthopedics": "Orthopedic",
    "ent": "Otolaryngology",
    "otolaryngology ent": "Otolaryngology",
    "otolaryngology_ent": "Otolaryngology",
    "otolaryngology": "Otolaryngology",
    "pulmonology": "Pulmonology",
    "research": "Research Studies",
    "research studies": "Research Studies",
    "seizure": "Seizures",
    "seizures": "Seizures",
    "treatment": "Treatments",
    "treatments": "Treatments",
    "exclude": "Excluded",
    "excluded": "Excluded",
    "exclusion": "Excluded",
}


@dataclass(frozen=True)
class LibrarySection:
    name: str
    description: str
    inclusion: str
    exclusion: str
    query: str


def cfc_clause() -> str:
    return '(cardiofaciocutaneous[All Fields] OR "cardio-facio-cutaneous"[All Fields] OR "CFC syndrome"[All Fields] OR "RASopathies"[MeSH Terms])'


SECTIONS: dict[str, LibrarySection] = {
    "Allergy and Immunology": LibrarySection(
        "Allergy and Immunology",
        "Immune dysregulation, hypersensitivity, eczema, allergic rhinitis, and recurrent infections in CFC and related RASopathies.",
        "Papers that directly investigate immune dysregulation, allergic disease, or inflammatory responses in individuals with confirmed or strongly suspected CFC syndrome, including RAS/MAPK mutations specifically linked to CFC.",
        "Studies focused solely on other RASopathies without CFC-specific data, generalized immunology papers without clinical relevance to CFC, or pathway studies lacking immune or allergy outcomes.",
        f'(("Hypersensitivity"[MeSH Terms] OR "Immunologic Deficiency Syndromes"[MeSH Terms] OR atopy[All Fields] OR allergic[All Fields] OR immune[All Fields]) AND {cfc_clause()})',
    ),
    "Cardiology": LibrarySection(
        "Cardiology",
        "Cardiac structure, hypertrophic cardiomyopathy, rhythm abnormalities, outflow obstruction, and contractile function in CFC.",
        "Research examining cardiac structure, function, or electrophysiology in CFC syndrome, including mechanistic studies using BRAF or MAPK mutations known to cause CFC.",
        "Cardiac studies on Noonan, Costello, or other RASopathies without a CFC subgroup, or broad cardiomyopathy papers not tied to CFC-relevant mutations.",
        f'(("Heart Defects, Congenital"[MeSH Terms] OR "Cardiomyopathies"[MeSH Terms] OR arrhythmia[All Fields] OR hypertrophic[All Fields]) AND {cfc_clause()})',
    ),
    "Dermatology": LibrarySection(
        "Dermatology",
        "Skin, hair, and nail findings in CFC, including eczema-like rashes, hyperkeratosis, keratosis pilaris, sparse or curly hair, and fragile nails.",
        "Studies describing skin, hair, or nail abnormalities in CFC syndrome, or mechanistic dermatologic work using CFC-associated RAS/MAPK mutations.",
        "Dermatology papers on other RASopathies without CFC-specific findings, or general dermatologic pathway studies not linked to CFC phenotypes.",
        f'(("Skin Manifestations"[MeSH Terms] OR "Skin Diseases, Genetic"[MeSH Terms] OR dermatology[All Fields] OR cutaneous[All Fields] OR hair[All Fields] OR nail[All Fields]) AND {cfc_clause()})',
    ),
    "Development and Behavior": LibrarySection(
        "Development and Behavior",
        "Cognitive, motor milestones, speech, behavioral, and growth-related developmental outcomes in CFC.",
        "Papers addressing motor milestones, cognition, speech/language, or behavioral phenotypes in individuals with CFC, including mechanistic work connecting RAS/MAPK dysregulation to developmental outcomes.",
        "Behavioral or developmental studies on other RASopathies without CFC representation, or broad motor-milestone research not tied to CFC-relevant mutations.",
        f'(("Developmental Disabilities"[MeSH Terms] OR "Behavioral Symptoms"[MeSH Terms] OR "Intellectual Disability"[MeSH Terms] OR cognition[All Fields] OR behavior[All Fields]) AND {cfc_clause()})',
    ),
    "Endocrinology": LibrarySection(
        "Endocrinology",
        "Hormonal, metabolic, and organ-related abnormalities associated with CFC syndrome and other RASopathies, including endocrine function, growth regulation, pubertal timing, gastrointestinal and renal organ development, feeding difficulties, failure to thrive, growth hormone deficiency, delayed or atypical puberty, renal anomalies, and electrolyte imbalances.",
        "Studies examining hormonal regulation, growth hormone function, pubertal development, metabolic abnormalities, or endocrine organ involvement in individuals with CFC; mechanistic work using CFC-associated mutations to model endocrine dysfunction.",
        "Endocrine studies focused solely on other RASopathies without CFC data; general metabolic or hormonal pathway papers not linked to CFC-specific clinical features.",
        f'(("Endocrine System Diseases"[MeSH Terms] OR hormone[All Fields] OR renal[All Fields] OR metabolic[All Fields] OR puberty[All Fields] OR pubertal[All Fields] OR "growth hormone"[All Fields] OR electrolyte[All Fields]) AND {cfc_clause()})',
    ),
    "Gastroenterology": LibrarySection(
        "Gastroenterology",
        "Feeding difficulties, reflux, constipation, motility disorders, nutrition, and failure to thrive in CFC.",
        "Studies investigating feeding difficulties, reflux, motility disorders, or gastrointestinal dysfunction in CFC syndrome, including mechanistic work using CFC-relevant RAS/MAPK mutations.",
        "GI studies on Noonan, Costello, or other RASopathies without CFC-specific findings, or general GI physiology papers not tied to CFC.",
        f'(("Gastrointestinal Diseases"[MeSH Terms] OR "Feeding and Eating Disorders"[MeSH Terms] OR reflux[All Fields] OR constipation[All Fields] OR motility[All Fields]) AND {cfc_clause()})',
    ),
    "General and Reviews": LibrarySection(
        "General and Reviews",
        "Foundational, overview, diagnostic, clinical-spectrum, guideline, and review papers about CFC within the RASopathy family.",
        "Foundational papers, clinical overviews, diagnostic criteria or CFC index papers, guidelines, prenatal diagnosis overviews, case reports or case series without a clearly specialty-specific new finding, adult CFC case reports or literature reviews, and RASopathy-wide reviews that meaningfully discuss CFC syndrome, its diagnostic criteria, phenotype spectrum, management, or its place within the RASopathy family.",
        "Broad RAS/MAPK, cancer, pathway, or RASopathy papers that only mention CFC in passing, only allude to CFC in a figure/table/reference, or do not provide CFC clinical, genetic, diagnostic, management, or meaningful comparison content.",
        f'({cfc_clause()} AND (Review[Publication Type] OR guideline[Publication Type] OR diagnosis[All Fields] OR phenotype[All Fields]))',
    ),
    "Genetics": LibrarySection(
        "Genetics",
        "Molecular etiology, mutation spectrum, genotype-phenotype correlation, and diagnostic genetics of CFC.",
        "Studies that present genetic analyses, mutation identification, or molecular characterization of CFC syndrome, including work that uses CFC-associated RAS/MAPK mutations to explore developmental mechanisms.",
        "Genetic studies focused solely on other RASopathies without CFC data, or general molecular pathway papers not directly linked to CFC-specific mutations or phenotypes.",
        f'(("Genetic Phenomena"[MeSH Terms] OR genetics[Subheading] OR mutation[All Fields] OR BRAF[All Fields] OR MAP2K1[All Fields] OR MAP2K2[All Fields] OR KRAS[All Fields]) AND {cfc_clause()})',
    ),
    "Growth": LibrarySection(
        "Growth",
        "Physical growth patterns, stature, skeletal maturation, and pubertal development across CFC syndrome and the broader Noonan-spectrum RASopathies, including growth hormone signaling, bone age progression, linear growth velocity, timing of puberty, short stature, delayed or disproportionate growth, growth hormone deficiency, and atypical pubertal onset.",
        "Studies addressing growth, stature, bone age, or pubertal development in CFC or Noonan-spectrum RASopathies; research using CFC-associated or Noonan-associated RAS/MAPK mutations to model growth hormone or pubertal regulation.",
        "Growth studies unrelated to RAS/MAPK signaling or lacking CFC/Noonan-spectrum relevance; general endocrinology papers without developmental or pubertal data tied to RASopathies.",
        '((growth[All Fields] OR stature[All Fields] OR "short stature"[All Fields] OR "bone age"[All Fields] OR "growth hormone"[All Fields] OR puberty[All Fields] OR pubertal[All Fields] OR "linear growth"[All Fields]) AND (cardiofaciocutaneous[All Fields] OR "cardio-facio-cutaneous"[All Fields] OR "CFC syndrome"[All Fields] OR "Noonan Syndrome"[MeSH Terms] OR "noonan spectrum"[All Fields] OR RASopathies[MeSH Terms]))',
    ),
    "Gynecology": LibrarySection(
        "Gynecology",
        "Reproductive, pubertal, menstrual, genital tract, and reproductive endocrine findings in CFC.",
        "Studies describing gynecologic, reproductive, or pubertal findings in individuals with CFC or using CFC-associated mutations to model reproductive endocrine function.",
        "Papers focused solely on other RASopathies without CFC data, or general reproductive endocrinology studies not tied to CFC-specific mutations or phenotypes.",
        f'(("Gynecology"[MeSH Terms] OR puberty[All Fields] OR menstrual[All Fields] OR reproductive[All Fields] OR ovarian[All Fields] OR uterine[All Fields]) AND {cfc_clause()})',
    ),
    "Neurology": LibrarySection(
        "Neurology",
        "Brain structure, hypotonia, motor delay, neuroimaging, coordination, and neurological physiology in CFC.",
        "Studies that document neurological findings, neuroimaging results, tone or motor abnormalities, or neurodevelopmental physiology in individuals with CFC, including mechanistic work using CFC-associated RAS/MAPK mutations.",
        "Neurologic studies focused solely on other RASopathies without CFC data, or general neuroscience papers not tied to CFC-specific phenotypes.",
        f'(("Nervous System Malformations"[MeSH Terms] OR "Neurodevelopmental Disorders"[MeSH Terms] OR hypotonia[All Fields] OR neuroimaging[All Fields] OR brain[All Fields]) AND {cfc_clause()})',
    ),
    "Oncology": LibrarySection(
        "Oncology",
        "Cancer risk, tumor formation, oncogenic RAS/MAPK mechanisms, and malignancy surveillance in CFC.",
        "Studies examining tumor development, cancer risk, or oncogenic mechanisms in individuals with CFC or using CFC-associated RAS/MAPK mutations to model malignancy.",
        "Oncology papers focused solely on other RASopathies without CFC data, or general cancer pathway studies not directly linked to CFC-specific mutations or phenotypes.",
        f'(("Neoplasms"[MeSH Terms] OR "Neoplastic Syndromes, Hereditary"[MeSH Terms] OR tumor[All Fields] OR cancer[All Fields] OR leukemia[All Fields]) AND {cfc_clause()})',
    ),
    "Ophthalmology": LibrarySection(
        "Ophthalmology",
        "Ocular structure, visual function, ptosis, strabismus, refractive errors, optic nerve findings, and corneal abnormalities in CFC.",
        "Research documenting ocular structure, function, or visual outcomes in CFC, or mechanistic studies using CFC-associated mutations to model eye development.",
        "Ophthalmology papers on other RASopathies without CFC representation, or general eye development studies not linked to CFC.",
        f'(("Eye Abnormalities"[MeSH Terms] OR "Eye Manifestations"[MeSH Terms] OR ophthalmology[All Fields] OR ocular[All Fields] OR strabismus[All Fields] OR nystagmus[All Fields]) AND {cfc_clause()})',
    ),
    "Orthopedic": LibrarySection(
        "Orthopedic",
        "Skeletal development, bone density, joint laxity, mineralization, posture, and mobility findings in CFC.",
        "Studies examining skeletal development, bone density, joint structure, or orthopedic manifestations in individuals with CFC, including mechanistic work using CFC-associated RAS/MAPK mutations.",
        "Orthopedic or bone metabolism studies focused solely on other RASopathies without CFC data, or general musculoskeletal research not tied to CFC-specific phenotypes.",
        f'(("Musculoskeletal Abnormalities"[MeSH Terms] OR skeletal[All Fields] OR bone[All Fields] OR orthopedic[All Fields] OR joint[All Fields]) AND {cfc_clause()})',
    ),
    "Otolaryngology": LibrarySection(
        "Otolaryngology",
        "Ear, nose, throat, hearing, airway, swallowing, sinus, laryngeal, and craniofacial ENT findings in CFC.",
        "Studies examining hearing, airway, swallowing, or ENT-related structural findings in CFC, including those using CFC-associated mutations to model craniofacial development.",
        "ENT studies on other RASopathies without CFC-specific data, or general airway/ENT research not tied to CFC phenotypes.",
        f'(("Otorhinolaryngologic Diseases"[MeSH Terms] OR hearing[All Fields] OR airway[All Fields] OR otitis[All Fields] OR laryngeal[All Fields] OR swallowing[All Fields]) AND {cfc_clause()})',
    ),
    "Pulmonology": LibrarySection(
        "Pulmonology",
        "Airway malformations, chronic lung disease, recurrent respiratory infections, aspiration, wheezing, cough, and sleep-disordered breathing in CFC.",
        "Research addressing lung function, airway structure, respiratory infections, or breathing disorders in individuals with CFC.",
        "Pulmonary studies on other RASopathies without CFC data, or general respiratory physiology papers not linked to CFC.",
        f'(("Respiratory Tract Diseases"[MeSH Terms] OR pulmonology[All Fields] OR respiratory[All Fields] OR lung[All Fields] OR apnea[All Fields] OR aspiration[All Fields]) AND {cfc_clause()})',
    ),
    "Research Studies": LibrarySection(
        "Research Studies",
        "Original experimental, clinical, molecular, developmental, and translational studies forming the scientific backbone for CFC pathophysiology.",
        "Studies presenting original data, experimental models, or clinical analyses directly involving individuals with CFC or using CFC-associated RAS/MAPK mutations to investigate developmental or molecular mechanisms.",
        "Research focused solely on other RASopathies without CFC representation, or general pathway studies not linked to CFC-specific clinical or genetic findings.",
        f'({cfc_clause()} NOT Review[Publication Type])',
    ),
    "Seizures": LibrarySection(
        "Seizures",
        "Epilepsy, seizure types, EEG findings, epileptic encephalopathy, medication response, and seizure mechanisms in CFC.",
        "Research focused on seizure types, EEG findings, epilepsy mechanisms, or seizure management in individuals with CFC, including studies using CFC-associated mutations to model epileptogenesis.",
        "Epilepsy studies on other RASopathies without CFC representation, or general seizure research not linked to CFC-specific genetic or clinical features.",
        f'(("Epilepsy"[MeSH Terms] OR "Seizures"[MeSH Terms] OR seizure[All Fields] OR epilepsy[All Fields] OR EEG[All Fields]) AND {cfc_clause()})',
    ),
    "Treatments": LibrarySection(
        "Treatments",
        "Therapeutic approaches, management strategies, targeted RAS/MAPK interventions, supportive therapies, and multidisciplinary care in CFC.",
        "Studies assessing treatments, interventions, or management strategies specifically for CFC, or therapeutic research using CFC-associated mutations to test targeted approaches.",
        "Treatment papers focused solely on other RASopathies without CFC data, or general therapeutic studies not applicable to CFC-specific clinical needs.",
        f'(("Therapeutics"[MeSH Terms] OR treatment[All Fields] OR therapy[All Fields] OR management[All Fields] OR inhibitor[All Fields] OR intervention[All Fields]) AND {cfc_clause()})',
    ),
    "Historical Articles": LibrarySection(
        "Historical Articles",
        "Historical, biography, and early descriptive papers relevant to the discovery and naming of CFC.",
        "Historical papers that clarify the origin, early recognition, diagnostic framing, or naming of CFC syndrome.",
        "Historical material that does not address CFC or CFC-relevant RASopathy classification.",
        f'({cfc_clause()} AND (Historical Article[Publication Type] OR Biography[Publication Type]))',
    ),
    "Conferences": LibrarySection(
        "Conferences",
        "Conference abstracts, proceedings, and congress papers that mention CFC.",
        "Conference material with explicit CFC clinical, genetic, mechanistic, or treatment relevance.",
        "Conference material about other RASopathies or general biology without CFC-specific information.",
        f'({cfc_clause()} AND Congresses[Publication Type])',
    ),
}

CATEGORY_CLASSIFICATION_GUIDANCE: dict[str, dict[str, list[str]]] = {
    "Allergy and Immunology": {
        "include_if": [
            "Primary objective is immune dysfunction, allergy, hypersensitivity, eczema/atopy, recurrent infection, autoimmunity, or inflammatory phenotype in CFC/RASopathy patients.",
            "Original cohort/case data include actual CFC participants or CFC-associated RAS/MAPK mutations with immune findings.",
        ],
        "exclude_if": [
            "Immune/allergy terms appear only in a phenotype list or background.",
            "Study is immunology-focused but lacks CFC data or CFC-relevant interpretation.",
            "Paper is a broad review without new immune-specific findings.",
        ],
        "strong_indicators": ["autoimmune", "autoantibodies", "immunodeficiency", "hypersensitivity", "atopy", "eczema", "recurrent infections", "inflammation"],
        "negative_examples": [
            "Broad RASopathy overview listing recurrent infections among many features -> General and Reviews, not Allergy and Immunology.",
            "Dermatology paper mentioning eczema as a skin finding -> Dermatology unless immune dysfunction is the main outcome.",
        ],
    },
    "Cardiology": {
        "include_if": [
            "Primary objective is cardiac disease, cardiac structure/function, rhythm, imaging, intervention, or cardiac outcomes.",
            "New cardiac phenotype, cardiac cohort, cardiac imaging, cardiomyopathy mechanism, arrhythmia, or surveillance outcome is central.",
        ],
        "exclude_if": [
            "Heart findings appear only in background or phenotype lists.",
            "Genetics paper only mentions cardiac findings while primary objective is mutation discovery.",
            "Review/guideline/management overview without new cardiac data.",
        ],
        "strong_indicators": ["hypertrophic cardiomyopathy", "pulmonary stenosis", "arrhythmia", "ventricle", "myocardium", "echocardiography", "cardiac MRI", "outflow obstruction"],
        "negative_examples": [
            "Clinical management of hypertrophic cardiomyopathy in CFC -> General and Reviews if it is a management review rather than original cardiac research.",
            "MAP2K1 variant report with cardiac findings -> Genetics if mutation discovery is the primary objective.",
        ],
    },
    "Dermatology": {
        "include_if": [
            "Primary objective is skin, hair, nails, cutaneous phenotype, dermatologic natural history, or skin-related mechanism.",
            "Paper presents new dermatologic findings, prospective dermatology data, or mechanistic cutaneous evidence.",
        ],
        "exclude_if": [
            "Skin findings appear only in broad phenotype background.",
            "Article is primarily genetics, review, or another organ system and only lists dermatologic symptoms.",
        ],
        "strong_indicators": ["cutaneous", "dermatologic", "skin", "hair", "nail", "keratosis", "eczema", "hyperkeratosis", "keratosis pilaris", "ulerythema"],
        "negative_examples": [
            "CFC clinical overview with a paragraph on skin -> General and Reviews, not Dermatology.",
            "Genotype paper listing sparse hair and keratosis -> Genetics unless dermatology is the main outcome.",
        ],
    },
    "Development and Behavior": {
        "include_if": [
            "Primary objective is cognition, behavior, speech/language, adaptive function, learning, or motor milestones.",
            "Study measures developmental or behavioral outcomes as a main endpoint.",
        ],
        "exclude_if": [
            "Developmental delay appears only as background.",
            "Primary outcome is seizures, neuroimaging, genetics, or growth rather than development/behavior.",
        ],
        "strong_indicators": ["behavior", "cognition", "adaptive", "speech", "language", "learning", "motor milestones", "developmental assessment", "intellectual disability"],
        "negative_examples": [
            "Epilepsy cohort with developmental delay listed -> Seizures if seizure phenotype is primary.",
            "Brain MRI paper with developmental context -> Neurology if imaging/brain findings are primary.",
        ],
    },
    "Endocrinology": {
        "include_if": [
            "Primary objective is endocrine, hormonal, metabolic, pubertal, renal/electrolyte, growth hormone physiology, or endocrine organ dysfunction.",
            "Mechanistic work links CFC/RAS/MAPK mutations to endocrine or metabolic dysfunction.",
        ],
        "exclude_if": [
            "Short stature or feeding difficulty is mentioned without endocrine focus.",
            "Physical growth trajectory is primary, which should usually be Growth.",
            "GI feeding/nutrition is primary, which should usually be Gastroenterology.",
        ],
        "strong_indicators": ["hormone", "growth hormone", "puberty", "pubertal", "metabolic", "renal", "electrolyte", "endocrine", "thyroid", "pituitary"],
        "negative_examples": [
            "Short stature cohort without endocrine testing -> Growth, not Endocrinology.",
            "Feeding and reflux study -> Gastroenterology unless endocrine/metabolic dysfunction is primary.",
        ],
    },
    "Gastroenterology": {
        "include_if": [
            "Primary objective is feeding, reflux, constipation, motility, nutrition, GI dysfunction, aspiration related to feeding, or failure to thrive as GI/nutrition issue.",
            "Paper presents new GI phenotype or management data specific to CFC/RASopathies.",
        ],
        "exclude_if": [
            "GI symptoms appear only in a phenotype list.",
            "Endocrine/metabolic or growth hormone physiology is primary.",
        ],
        "strong_indicators": ["feeding", "reflux", "constipation", "motility", "gastrointestinal", "nutrition", "failure to thrive", "swallowing", "aspiration"],
        "negative_examples": [
            "Broad clinical features paper listing reflux -> General and Reviews unless GI is the main objective.",
            "Growth hormone/puberty paper mentioning feeding -> Endocrinology or Growth, not Gastroenterology.",
        ],
    },
    "General and Reviews": {
        "include_if": [
            "Article type is review, clinical guideline, diagnostic paper, CFC index, management paper, broad clinical overview, broad phenotype-spectrum paper, or patient/clinical summary.",
            "Case report/case series has no clearly specialty-specific new finding and mainly contributes broad clinical characterization.",
            "Adult CFC case reports, adult CFC literature reviews, or adult-with-CFC clinical summaries belong here unless the paper presents a clearly novel specialty-specific finding.",
        ],
        "exclude_if": [
            "Original research has a clear specialty-specific primary outcome.",
            "Mutation discovery/genotype-phenotype paper is primary Genetics.",
            "Treatment landscape paper is primary Treatments.",
            "General background appears before a more specific study objective.",
        ],
        "strong_indicators": ["review", "literature review", "guideline", "diagnostic criteria", "diagnosis", "index", "overview", "management", "clinical features", "phenotype spectrum", "consensus", "adult"],
        "negative_examples": [
            "Dermatology case series with background review -> Dermatology.",
            "Genetic variant discovery paper with broad introduction -> Genetics.",
            "Treatment-focused review -> Treatments, with organ system only if substantial.",
        ],
    },
    "Genetics": {
        "include_if": [
            "Primary objective is mutation discovery, variant validation, genotype-phenotype correlation, inheritance, molecular diagnosis, or variant function.",
            "CFC-associated BRAF, MAP2K1, MAP2K2, KRAS, or RAS/MAPK mutation biology is the central scientific question.",
        ],
        "exclude_if": [
            "Genes are mentioned only as background for a specialty phenotype.",
            "Paper is broad review/guideline rather than original genetic analysis.",
        ],
        "strong_indicators": ["BRAF", "MAP2K1", "MAP2K2", "KRAS", "variant", "mutation", "genotype", "phenotype correlation", "de novo", "molecular diagnosis"],
        "negative_examples": [
            "Dermatology study noting patients are HRAS/BRAF positive -> Dermatology if skin outcomes are primary.",
            "Clinical management guideline mentioning genes -> General and Reviews.",
        ],
    },
    "Growth": {
        "include_if": [
            "Primary objective is stature, linear growth, growth velocity, bone age, pubertal growth, growth hormone treatment/outcomes, or growth trajectory.",
            "Noonan-spectrum growth study is included when it informs shared CFC/RASopathy growth mechanisms.",
        ],
        "exclude_if": [
            "Short stature appears only in phenotype list.",
            "Feeding/failure to thrive is primary GI/nutrition issue.",
            "Endocrine physiology rather than growth trajectory is primary.",
        ],
        "strong_indicators": ["growth", "stature", "short stature", "linear growth", "growth velocity", "bone age", "puberty", "pubertal", "growth hormone"],
        "negative_examples": [
            "Broad phenotype paper listing short stature -> General and Reviews.",
            "Feeding difficulty/failure-to-thrive study -> Gastroenterology unless growth trajectory is the main outcome.",
        ],
    },
    "Gynecology": {
        "include_if": [
            "Primary objective is reproductive, menstrual, genital tract, ovarian/uterine, pubertal gynecologic, or reproductive endocrine phenotype.",
            "CFC/RASopathy patient data or CFC-associated mutation model is present.",
        ],
        "exclude_if": [
            "Puberty is discussed only as general growth/endocrine background.",
            "Study is endocrine but not gynecologic/reproductive.",
        ],
        "strong_indicators": ["menstrual", "reproductive", "ovarian", "uterine", "gynecologic", "genital", "puberty", "pubertal development"],
        "negative_examples": [
            "Growth/puberty paper without gynecologic focus -> Growth or Endocrinology.",
            "Broad review listing delayed puberty -> General and Reviews.",
        ],
    },
    "Neurology": {
        "include_if": [
            "Primary objective is brain structure, hypotonia, motor delay, neuroimaging, neurologic physiology, cognition related to neurologic findings, or white matter/brain abnormalities.",
            "Neurologic phenotype is the main outcome and seizure is not the primary focus.",
        ],
        "exclude_if": [
            "Seizures/EEG/epilepsy are the primary outcome, which should be Seizures.",
            "Development/behavior scales are primary without neurologic findings, which may be Development and Behavior.",
        ],
        "strong_indicators": ["hypotonia", "motor delay", "MRI", "neuroimaging", "brain", "white matter", "cognition", "neurologic", "coordination"],
        "negative_examples": [
            "EEG/seizure phenotype paper -> Seizures, not Neurology.",
            "Behavioral assessment paper without neurologic findings -> Development and Behavior.",
        ],
    },
    "Oncology": {
        "include_if": [
            "Primary objective is tumor development, cancer risk, malignancy surveillance, oncogenic mechanism, leukemia, melanoma, or neoplasm in CFC/RASopathy context.",
            "CFC-associated mutation is used to model malignancy or oncogenic RAS/MAPK biology.",
        ],
        "exclude_if": [
            "Somatic cancer mutation paper does not discuss CFC clinically or mechanistically.",
            "Cancer biology mentions BRAF/MAPK but lacks CFC relevance.",
        ],
        "strong_indicators": ["tumor", "cancer", "neoplasm", "malignancy", "leukemia", "melanoma", "oncogenic", "surveillance"],
        "negative_examples": [
            "RASopathy gene mutations in melanoma without substantive CFC discussion -> Excluded.",
            "General BRAF cancer pathway paper -> Excluded unless CFC-specific interpretation exists.",
        ],
    },
    "Ophthalmology": {
        "include_if": [
            "Primary objective is ocular structure, vision, strabismus, nystagmus, ptosis, refractive error, optic nerve, retina, cornea, or eye development.",
            "Paper presents ocular phenotype/outcomes in CFC/RASopathy patients or CFC mutation model.",
        ],
        "exclude_if": [
            "Eye findings appear only in a phenotype list.",
            "Broad clinical review has no new ocular finding.",
        ],
        "strong_indicators": ["ocular", "ophthalmology", "vision", "strabismus", "nystagmus", "ptosis", "refractive", "optic nerve", "retina", "cornea"],
        "negative_examples": [
            "CFC overview listing strabismus -> General and Reviews.",
            "Genetic case report mentioning eye findings -> Genetics if variant discovery is primary.",
        ],
    },
    "Orthopedic": {
        "include_if": [
            "Primary objective is skeletal, bone, joint, mineralization, posture, scoliosis, mobility, or musculoskeletal phenotype.",
            "Paper presents orthopedic outcomes or mechanistic skeletal data.",
        ],
        "exclude_if": [
            "Musculoskeletal findings appear only in background or broad phenotype list.",
            "Growth/stature is primary, which should usually be Growth.",
        ],
        "strong_indicators": ["skeletal", "bone", "joint", "orthopedic", "scoliosis", "musculoskeletal", "bone density", "mineralization", "mobility"],
        "negative_examples": [
            "Short stature paper with skeletal maturation -> Growth if growth trajectory is primary.",
            "Broad phenotype overview listing scoliosis -> General and Reviews.",
        ],
    },
    "Otolaryngology": {
        "include_if": [
            "Primary objective is ear, nose, throat, hearing, airway, swallowing, sinus, laryngeal, craniofacial ENT, otitis, or ENT structural finding.",
            "Paper presents ENT/hearing/airway outcomes in CFC/RASopathy patients or CFC mutation model.",
        ],
        "exclude_if": [
            "ENT findings appear only in a broad phenotype list.",
            "Respiratory/lung disease is primary, which should be Pulmonology.",
            "Feeding/GI swallowing/nutrition is primary, which may be Gastroenterology.",
        ],
        "strong_indicators": ["hearing", "airway", "otitis", "laryngeal", "ENT", "otorhinolaryngologic", "swallowing", "sinus", "craniofacial"],
        "negative_examples": [
            "Pulmonary aspiration paper -> Pulmonology or Gastroenterology depending on primary objective.",
            "Broad clinical review listing hearing loss -> General and Reviews.",
        ],
    },
    "Pulmonology": {
        "include_if": [
            "Primary objective is lung disease, airway malformation, respiratory infection, aspiration, wheeze, cough, sleep-disordered breathing, or pulmonary function.",
            "Paper presents respiratory outcomes or pulmonary management data in CFC/RASopathy patients.",
        ],
        "exclude_if": [
            "Respiratory symptoms appear only in phenotype list.",
            "ENT airway/hearing focus is primary, which should be Otolaryngology.",
        ],
        "strong_indicators": ["respiratory", "lung", "pulmonary", "airway", "apnea", "aspiration", "wheezing", "sleep-disordered breathing", "infection"],
        "negative_examples": [
            "ENT airway structural paper -> Otolaryngology.",
            "Broad phenotype review listing respiratory infections -> General and Reviews.",
        ],
    },
    "Research Studies": {
        "include_if": [
            "Primary objective is original experimental, translational, molecular, developmental, or model-system research that is CFC-relevant but not better captured by a specialty category.",
            "Study uses CFC-associated RAS/MAPK mutations to investigate disease mechanism without a dominant organ-system category.",
        ],
        "exclude_if": [
            "A clearer specialty category exists.",
            "Review/guideline/overview rather than original research.",
            "Broad pathway work lacks CFC-specific mutation/model interpretation.",
        ],
        "strong_indicators": ["model", "mechanism", "experimental", "translational", "RAS/MAPK", "pathophysiology", "developmental mechanism", "cellular"],
        "negative_examples": [
            "Original cardiac mechanism study -> Cardiology if cardiac outcome is primary.",
            "Original genetic variant study -> Genetics.",
        ],
    },
    "Seizures": {
        "include_if": [
            "Primary objective is epilepsy, seizure phenotype, EEG, epileptic encephalopathy, antiseizure medication, or seizure mechanism.",
            "Paper presents seizure/EEG outcomes or seizure management in CFC/RASopathy patients or CFC mutation model.",
        ],
        "exclude_if": [
            "Seizures appear only in neurologic symptom list.",
            "Broader neuroimaging/hypotonia/cognition is primary, which should be Neurology or Development and Behavior.",
        ],
        "strong_indicators": ["epilepsy", "seizure", "infantile spasms", "EEG", "antiseizure medication", "epileptic encephalopathy", "status epilepticus"],
        "negative_examples": [
            "Neurologic phenotype paper listing seizures among many findings -> Neurology unless seizure/EEG is primary.",
            "Broad clinical review listing epilepsy -> General and Reviews.",
        ],
    },
    "Treatments": {
        "include_if": [
            "Primary objective is therapy, intervention, medication, surgery, treatment response, treatment strategy, therapeutic landscape, targeted RAS/MAPK intervention, or multidisciplinary management strategy.",
            "Article evaluates or reviews treatment options as its central purpose.",
        ],
        "exclude_if": [
            "Treatment is mentioned only in discussion, follow-up, supportive care, or background.",
            "Original specialty phenotype/genotype paper only briefly mentions management.",
            "General clinical overview is broad rather than treatment-focused.",
        ],
        "strong_indicators": ["treatment", "therapy", "therapeutic", "intervention", "medication", "surgery", "MEK inhibitor", "management strategy", "treatment response", "therapeutic landscape"],
        "negative_examples": [
            "Cardiac cohort mentioning management in discussion -> Cardiology.",
            "Broad CFC clinical guideline covering many topics -> General and Reviews unless treatment is the main objective.",
        ],
    },
}

ARTICLE_TYPES = [
    "Original research",
    "Case report",
    "Case series",
    "Review",
    "Clinical guideline",
    "Diagnostic paper",
    "Management paper",
]

CATEGORY_PRIORITY_RULES = [
    "If seizure/epilepsy/EEG is the primary outcome, choose Seizures rather than Neurology.",
    "If mutation discovery, variant validation, or genotype-phenotype correlation is the primary objective, choose Genetics even if organ findings are described.",
    "If therapeutic intervention or therapeutic landscape is the primary objective, choose Treatments; add an organ-system category only when at least about 30% of the paper focuses on that organ system.",
    "If the article is a review, clinical guideline, diagnostic paper, management paper, or broad clinical overview, choose General and Reviews unless it presents substantial new specialty-specific findings.",
    "If the article is an adult CFC case report, adult-with-CFC clinical summary, or adult CFC literature review, choose General and Reviews unless it presents a clearly novel specialty-specific finding.",
    "If skin/hair/nail findings are the primary objective, choose Dermatology even when CFC genetics are mentioned.",
    "If cardiac disease/function/imaging is the primary objective, choose Cardiology unless the article is primarily a broad review/guideline.",
    "If growth trajectory, stature, bone age, or pubertal growth is the primary objective, choose Growth; if endocrine physiology is primary, choose Endocrinology.",
    "If no specialty category clearly dominates original CFC-relevant mechanistic research, choose Research Studies.",
]


SECTION_ALIASES = {
    "ENT": "Otolaryngology",
    "Otolaryngology ENT": "Otolaryngology",
    "Ophtalmology": "Ophthalmology",
}


FINAL_COLUMNS = [
    "Primary_Category",
    "Matched_Categories",
    "Reviewer_Decision",
    "PMID",
    "Title",
    "Authors",
    "Journal",
    "Publication_Year",
    "Publication_Date",
    "DOI",
    "Abstract",
    "Suggested_Labels",
    "Eligibility_Decision",
    "Eligibility_Rationale",
    "History_Match",
    "History_Decision",
    "History_Source",
    "Section_Inclusion_Criteria",
    "Section_Exclusion_Criteria",
    "Found_in_Zotero",
    "Review_Status",
    "Reviewer_Notes",
    "PubMed_URL",
    "Last_Seen",
]

REVIEW_COMPARISON_YEARS = (2017, 2022)

REVIEW_COMPARISON_EXPORT_COLUMNS = [
    ("PMID", "PubMed ID"),
    ("Title", "Article Title"),
    ("Publication_Date", "Date"),
    ("System_Relevance_Decision", "Relevance"),
    ("Google_Sheets_Found", "Found in Sheets?"),
    ("Zotero_Found", "Found in Zotero?"),
    ("Human_OpenAI_Match", "Match?"),
    ("OpenAI_Screening_Display", "OpenAI Screening In or Out"),
    ("Zotero_Category_Display", "If in Zotero - Category"),
    ("OpenAI_Assigned_Category_Display", "OpenAI Assigned Category"),
]


def normalize_category(name: str) -> str:
    cleaned = name.strip()
    return SECTION_ALIASES.get(cleaned, cleaned)


def load_env_file(path: Path = Path(".env")) -> None:
    """Load simple KEY=VALUE entries from .env without overriding existing variables."""
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def normalize_pmid(value: object) -> str:
    if value is None:
        return ""
    match = re.search(r"\b(\d{4,})\b", str(value))
    return match.group(1) if match else ""


def normalize_doi(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("_", " ")
    text = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", text)
    text = re.sub(r"^doi:\s*", "", text)
    return text.strip().rstrip(".")


def screening_direction(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    screen_in_terms = (
        "screen in",
        "screened in",
        "include",
        "included",
        "approved",
        "relevant",
        "yes",
    )
    screen_out_terms = (
        "screen out",
        "screened out",
        "exclude",
        "excluded",
        "not approved",
        "not relevant",
        "no",
    )
    if any(term in text for term in screen_out_terms):
        return "out"
    if any(term in text for term in screen_in_terms):
        return "in"
    return ""


def split_category_labels(value: object) -> set[str]:
    if value is None or pd.isna(value):
        return set()
    text = str(value).strip().lower()
    if not text:
        return set()
    protected = {
        "allergy and immunology": "allergy_immunology",
        "development and behavior": "development_behavior",
        "general and reviews": "general_reviews",
    }
    for phrase, token in protected.items():
        text = text.replace(phrase, token)
    text = re.sub(r"\badditional\s*:\s*", ";", text)
    text = text.replace("&", ";")
    text = re.sub(r"\band\b", ";", text)
    text = text.replace(",", ";").replace("/", ";").replace("|", ";")
    for phrase, token in protected.items():
        text = text.replace(token, phrase)
    parts = [re.sub(r"\s+", " ", part).strip(" .;:?") for part in text.split(";")]
    categories: set[str] = set()
    for part in parts:
        if not part:
            continue
        if part in CATEGORY_ALIASES:
            categories.add(CATEGORY_ALIASES[part])
            continue
        for alias, category in CATEGORY_ALIASES.items():
            if re.search(rf"\b{re.escape(alias)}\b", part):
                categories.add(category)
    return categories


def comparison_category_text(primary: object, secondary: object = "") -> str:
    categories = split_category_labels(primary)
    categories.update(split_category_labels(secondary))
    ordered = [
        category
        for category in SECTIONS
        if category in categories and category not in EXCLUDED_COMPARISON_FOLDERS
    ]
    if not ordered:
        return ""
    primary_category = ordered[0]
    additional = [category for category in ordered[1:] if category != primary_category]
    if additional:
        return f"{primary_category}; Additional: {additional[0]}"
    return primary_category


def require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def load_runtime_dependencies() -> None:
    global pd, Entrez, zotero, SentenceTransformer, util

    try:
        import pandas as pandas_module
        from Bio import Entrez as entrez_module
        from pyzotero import zotero as zotero_module
        from sentence_transformers import SentenceTransformer as sentence_transformer_class
        from sentence_transformers import util as sentence_transformer_util
    except ModuleNotFoundError as exc:
        missing = exc.name or "a required package"
        raise RuntimeError(
            f"Missing Python package '{missing}'. Install dependencies with: python -m pip install -r requirements.txt"
        ) from exc

    pd = pandas_module
    Entrez = entrez_module
    zotero = zotero_module
    SentenceTransformer = sentence_transformer_class
    util = sentence_transformer_util


def materialize_history_source(source: str) -> Path:
    parsed = urlparse(source)
    if parsed.scheme in {"http", "https"}:
        match = re.search(r"/spreadsheets/d/([^/]+)", source)
        download_url = source
        suffix = ".xlsx"
        if match:
            download_url = f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"
        target = Path(tempfile.gettempdir()) / f"cfc_screening_history_{int(time.time())}{suffix}"
        urlretrieve(download_url, target)
        return target
    return Path(source)


def find_column(df: DataFrame, candidates: tuple[str, ...]) -> str | None:
    normalized = {normalize_text(column): column for column in df.columns}
    for candidate in candidates:
        key = normalize_text(candidate)
        if key in normalized:
            return normalized[key]
    return None


def load_screening_history(source: str | None) -> dict[str, dict[str, dict[str, str]]]:
    empty = {"pmid": {}, "doi": {}, "title": {}}
    if not source:
        return empty

    path = materialize_history_source(source)
    if not path.exists():
        raise RuntimeError(f"Screening history file was not found: {path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
        frames = [frame.assign(_history_sheet=name) for name, frame in sheets.items()]
        history_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    elif suffix == ".csv":
        history_df = pd.read_csv(path, dtype=str)
        history_df["_history_sheet"] = path.name
    elif suffix == ".tsv":
        history_df = pd.read_csv(path, sep="\t", dtype=str)
        history_df["_history_sheet"] = path.name
    else:
        raise RuntimeError("Screening history must be an .xlsx, .xls, .csv, .tsv, or accessible Google Sheets URL.")

    if history_df.empty:
        return empty

    pmid_col = find_column(history_df, ("PMID", "PubMed ID", "PubMed_ID", "PMID Number"))
    doi_col = find_column(history_df, ("DOI", "doi", "Digital Object Identifier"))
    title_col = find_column(history_df, ("Title", "Article Title", "Article_Title"))
    decision_col = find_column(history_df, ("Eligibility_Decision", "Eligibility Decision", "Decision", "Review_Status", "Review Status"))
    status_col = find_column(history_df, ("Review_Status", "Review Status", "Status"))
    category_col = find_column(history_df, ("Category", "Primary_Category", "Primary Category", "Folder", "Zotero Folder", "Suggested_Labels", "Suggested Labels"))
    notes_col = find_column(history_df, ("Notes", "Reviewer_Notes", "Reviewer Notes", "Rationale", "Comment", "Comments"))

    history = {"pmid": {}, "doi": {}, "title": {}}
    for idx, row in history_df.iterrows():
        pmid = normalize_pmid(row.get(pmid_col)) if pmid_col else ""
        doi = normalize_doi(row.get(doi_col)) if doi_col else ""
        title = normalize_text(row.get(title_col)) if title_col else ""
        decision = str(row.get(decision_col, "") or row.get(status_col, "") or "").strip()
        category = str(row.get(category_col, "") or "").strip() if category_col else ""
        notes = str(row.get(notes_col, "") or "").strip() if notes_col else ""
        source_label = f"{path.name}:{row.get('_history_sheet', 'Sheet')}:{idx + 2}"
        record = {"decision": decision, "category": category, "notes": notes, "source": source_label, "doi": doi}
        if pmid:
            history["pmid"][pmid] = record
        if doi:
            history["doi"][doi] = record
        if title:
            history["title"][title] = record
    return history


def lookup_history_match(pmid: str, title: str, history: dict[str, dict[str, dict[str, str]]], doi: str = "") -> dict[str, str] | None:
    if pmid and pmid in history.get("pmid", {}):
        return history["pmid"][pmid]
    normalized_doi = normalize_doi(doi)
    if normalized_doi and normalized_doi in history.get("doi", {}):
        return history["doi"][normalized_doi]
    normalized_title = normalize_text(title)
    if normalized_title and normalized_title in history.get("title", {}):
        return history["title"][normalized_title]
    return None


def extract_pmid_from_zotero_extra(extra: str | None) -> str | None:
    if not extra:
        return None
    match = re.search(r"\bPMID:\s*(\d+)\b", extra, flags=re.IGNORECASE)
    return match.group(1) if match else None


def zotero_item_record(item: dict, data: dict, collections: dict[str, str]) -> dict[str, str]:
    collection_keys = data.get("collections") or []
    folder_names = [collections.get(key, key) for key in collection_keys if collections.get(key, key)]
    current_category = ", ".join(folder_names) if folder_names else "unfiled"
    return {
        "found": "Yes",
        "category": current_category,
        "file_status": "filed" if folder_names else "unfiled",
        "item_key": item.get("key", ""),
        "url": data.get("url", ""),
        "pmid": extract_pmid_from_zotero_extra(data.get("extra")) or "",
        "doi": normalize_doi(data.get("DOI", "")),
        "title": normalize_text(data.get("title", "")),
    }


def fetch_zotero_pmids(group_id: str, api_key: str) -> set[str]:
    zot = zotero.Zotero(group_id, "group", api_key)
    items = zot.everything(zot.items())
    pmids = set()
    for item in items:
        data = item.get("data", {})
        pmid = extract_pmid_from_zotero_extra(data.get("extra"))
        if pmid:
            pmids.add(pmid)
    return pmids


def fetch_zotero_index(group_id: str, api_key: str) -> dict[str, dict[str, dict[str, str]]]:
    zot = zotero.Zotero(group_id, "group", api_key)
    collections = {collection["key"]: collection["data"].get("name", "") for collection in zot.everything(zot.collections())}
    items = zot.everything(zot.items())
    index: dict[str, dict[str, dict[str, str]]] = {"pmid": {}, "doi": {}, "title": {}}
    for item in items:
        data = item.get("data", {})
        record = zotero_item_record(item, data, collections)
        if record["pmid"]:
            index["pmid"][record["pmid"]] = record
        if record["doi"]:
            index["doi"][record["doi"]] = record
        if record["title"]:
            index["title"][record["title"]] = record
    return index


def build_zotero_assignment_examples(group_id: str, api_key: str) -> str:
    zot = zotero.Zotero(group_id, "group", api_key)
    collections = {collection["key"]: collection["data"].get("name", "") for collection in zot.everything(zot.collections())}
    items = zot.everything(zot.items())
    examples: dict[str, list[str]] = {}

    for item in items:
        data = item.get("data", {})
        if data.get("itemType") in {"attachment", "note", "annotation"}:
            continue
        title = re.sub(r"\s+", " ", str(data.get("title", "") or "")).strip()
        if not title:
            continue
        abstract = re.sub(r"\s+", " ", str(data.get("abstractNote", "") or "")).strip()
        collection_names = [collections.get(key, key) for key in data.get("collections", [])]
        category_labels: set[str] = set()
        for collection_name in collection_names:
            category_labels.update(split_category_labels(collection_name))
        if not category_labels:
            continue

        for category in sorted(category_labels):
            examples.setdefault(category, [])
            if len(examples[category]) >= MAX_ZOTERO_EXAMPLES_PER_CATEGORY:
                continue
            example_text = f"- {title}"
            if abstract:
                example_text += f" | Abstract clue: {abstract[:450]}"
            examples[category].append(example_text)

    lines = [
        "Use these existing Zotero folder assignments as examples of Lexi/Rich's categorization style.",
        "These examples are guidance, not automatic truth: still apply the inclusion/exclusion criteria and full-text evidence.",
        "If an example is in Excluded, learn the screen-out pattern rather than treating Excluded as a clinical category.",
    ]
    for category in list(SECTIONS) + ["Excluded"]:
        category_examples = examples.get(category, [])
        if not category_examples:
            continue
        lines.append(f"\n{category} examples:")
        lines.extend(category_examples)
    return "\n".join(lines)


def lookup_zotero_match(
    pmid: str,
    title: str,
    doi: str,
    zotero_index: dict[str, dict[str, dict[str, str]]],
) -> dict[str, str] | None:
    if pmid and pmid in zotero_index.get("pmid", {}):
        return zotero_index["pmid"][pmid]
    normalized_doi = normalize_doi(doi)
    if normalized_doi and normalized_doi in zotero_index.get("doi", {}):
        return zotero_index["doi"][normalized_doi]
    normalized_title = normalize_text(title)
    if normalized_title and normalized_title in zotero_index.get("title", {}):
        return zotero_index["title"][normalized_title]
    return None


def zotero_api_get_json(url: str, api_key: str) -> dict | list:
    request = Request(
        url,
        headers={"Zotero-API-Key": api_key, "Accept": "application/json"},
    )
    try:
        with urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception:
        return {}


def zotero_fulltext_for_item(group_id: str, api_key: str, item_key: str) -> tuple[str, str]:
    if not item_key:
        return "", ""

    encoded_key = quote(item_key)
    base = f"https://api.zotero.org/groups/{group_id}/items/{encoded_key}"
    texts = []
    sources = []

    item_fulltext = zotero_api_get_json(f"{base}/fulltext", api_key)
    if isinstance(item_fulltext, dict) and item_fulltext.get("content"):
        texts.append(str(item_fulltext.get("content", "")))
        sources.append("item")

    children = zotero_api_get_json(f"{base}/children?limit=100", api_key)
    if isinstance(children, list):
        for child in children:
            child_data = child.get("data", {})
            if child_data.get("itemType") != "attachment":
                continue
            child_key = child.get("key") or child_data.get("key")
            if not child_key:
                continue
            attachment_fulltext = zotero_api_get_json(
                f"https://api.zotero.org/groups/{group_id}/items/{quote(child_key)}/fulltext",
                api_key,
            )
            if isinstance(attachment_fulltext, dict) and attachment_fulltext.get("content"):
                texts.append(str(attachment_fulltext.get("content", "")))
                sources.append(f"attachment:{child_key}")

    combined = "\n\n".join(texts)
    combined = re.sub(r"\s+", " ", combined).strip()
    return combined, "; ".join(sources)


def add_zotero_fulltext_columns(df: DataFrame, group_id: str, api_key: str) -> DataFrame:
    output = df.copy()
    output["Zotero_Full_Text_Found"] = "No"
    output["Zotero_Full_Text_Source"] = ""
    output["Zotero_Full_Text"] = ""

    if output.empty or "Zotero_Item_Key" not in output:
        return output

    for idx, row in output.iterrows():
        if str(row.get("Zotero_Found", "")).strip() != "Yes":
            continue
        item_key = str(row.get("Zotero_Item_Key", "") or "").strip()
        if not item_key:
            continue
        full_text, source = zotero_fulltext_for_item(group_id, api_key, item_key)
        time.sleep(0.25)
        if full_text:
            output.at[idx, "Zotero_Full_Text_Found"] = "Yes"
            output.at[idx, "Zotero_Full_Text_Source"] = source
            output.at[idx, "Zotero_Full_Text"] = full_text
    return output


def publication_date_filter(from_year: int | None, to_year: int | None) -> tuple[str | None, str | None]:
    if not from_year and not to_year:
        return None, None
    current_year = date.today().year
    if from_year and (from_year < 1800 or from_year > current_year):
        raise RuntimeError(f"--from-year must be between 1800 and {current_year}.")
    if to_year and (to_year < 1800 or to_year > current_year):
        raise RuntimeError(f"--to-year must be between 1800 and {current_year}.")
    from_year = from_year or 1800
    to_year = to_year or current_year
    if from_year > to_year:
        raise RuntimeError("--from-year must be earlier than or equal to --to-year.")
    return f"{from_year}/01/01", f"{to_year}/12/31"


def format_year_filter(from_year: int | None, to_year: int | None) -> str:
    if from_year and to_year:
        return f"{from_year}-{to_year}"
    if from_year:
        return f"{from_year} onward"
    if to_year:
        return f"through {to_year}"
    return "all years"


def search_pubmed(query: str, retmax: int, from_year: int | None, to_year: int | None) -> list[str]:
    mindate, maxdate = publication_date_filter(from_year, to_year)
    search_kwargs = {
        "db": "pubmed",
        "term": query,
        "retmax": str(retmax),
        "sort": "pub+date",
    }
    if mindate and maxdate:
        search_kwargs.update({"datetype": "pdat", "mindate": mindate, "maxdate": maxdate})
    record = entrez_call_with_retry(lambda: Entrez.read(Entrez.esearch(**search_kwargs)), "PubMed search")
    return [str(pmid) for pmid in record.get("IdList", [])]


def entrez_call_with_retry(operation, label: str, attempts: int = 5):
    for attempt in range(1, attempts + 1):
        try:
            return operation()
        except HTTPError as exc:
            retryable = exc.code in {429, 500, 502, 503, 504}
            if not retryable or attempt == attempts:
                raise
            wait_seconds = min(60, 8 * attempt)
            print(f"{label} was rate-limited by PubMed. Waiting {wait_seconds} seconds, then retrying...")
            time.sleep(wait_seconds)
        except URLError:
            if attempt == attempts:
                raise
            wait_seconds = min(60, 5 * attempt)
            print(f"{label} had a temporary network problem. Waiting {wait_seconds} seconds, then retrying...")
            time.sleep(wait_seconds)


def chunks(values: list[str], size: int) -> Iterable[list[str]]:
    for index in range(0, len(values), size):
        yield values[index : index + size]


def fetch_pubmed_records(pmids: list[str], batch_size: int = 200) -> list[dict]:
    records: list[dict] = []
    for batch in chunks(pmids, batch_size):
        data = entrez_call_with_retry(
            lambda: Entrez.read(Entrez.efetch(db="pubmed", id=",".join(batch), retmode="xml")),
            "PubMed record fetch",
        )
        records.extend(data.get("PubmedArticle", []))
    return records


def stringify_abstract(article: dict) -> str:
    abstract = article.get("MedlineCitation", {}).get("Article", {}).get("Abstract", {})
    parts = abstract.get("AbstractText", [])
    values = []
    for part in parts:
        label = getattr(part, "attributes", {}).get("Label")
        text = str(part)
        values.append(f"{label}: {text}" if label else text)
    return " ".join(values).strip()


def article_date(article: dict) -> tuple[str, str]:
    pubmed_data = article.get("PubmedData", {})
    history = pubmed_data.get("History", [])
    for event in history:
        if event.attributes.get("PubStatus") == "pubmed":
            year = str(event.get("Year", ""))
            month = str(event.get("Month", "01")).zfill(2)
            day = str(event.get("Day", "01")).zfill(2)
            return year, f"{year}-{month}-{day}"
    journal_issue = article.get("MedlineCitation", {}).get("Article", {}).get("Journal", {}).get("JournalIssue", {})
    pub_date = journal_issue.get("PubDate", {})
    year = str(pub_date.get("Year", ""))
    return year, year


def doi_from_article(article: dict) -> str:
    ids = article.get("PubmedData", {}).get("ArticleIdList", [])
    for article_id in ids:
        if article_id.attributes.get("IdType") == "doi":
            return str(article_id)
    return ""


def parse_pubmed_article(
    article: dict,
    section: LibrarySection,
    zotero_pmids: set[str],
    screening_history: dict[str, dict[str, dict[str, str]]] | None = None,
) -> dict:
    medline = article.get("MedlineCitation", {})
    article_data = medline.get("Article", {})
    pmid = str(medline.get("PMID", ""))
    title = str(article_data.get("ArticleTitle", "")).strip()
    abstract = stringify_abstract(article)
    authors = []
    for author in article_data.get("AuthorList", []):
        name = f"{author.get('LastName', '')} {author.get('Initials', '')}".strip()
        if name:
            authors.append(name)
    year, publication_date = article_date(article)
    journal = str(article_data.get("Journal", {}).get("Title", ""))
    doi = doi_from_article(article)
    history_match = lookup_history_match(pmid, title, screening_history or {"pmid": {}, "doi": {}, "title": {}}, doi)
    eligibility, rationale = screen_article(title, abstract, section)
    found_in_zotero = pmid in zotero_pmids
    return {
        "PMID": pmid,
        "Title": title,
        "Authors": ", ".join(authors) if authors else "N/A",
        "Journal": journal,
        "Publication_Year": year,
        "Publication_Date": publication_date,
        "DOI": doi,
        "Abstract": abstract,
        "Primary_Category": section.name,
        "Matched_Categories": section.name,
        "Suggested_Labels": "",
        "Eligibility_Decision": eligibility,
        "Eligibility_Rationale": rationale,
        "History_Match": bool(history_match),
        "History_Decision": history_match.get("decision", "") if history_match else "",
        "History_Source": history_match.get("source", "") if history_match else "",
        "Section_Inclusion_Criteria": section.inclusion,
        "Section_Exclusion_Criteria": section.exclusion,
        "Found_in_Zotero": found_in_zotero,
        "Reviewer_Decision": "Needs reviewed",
        "Review_Status": "Already in Zotero" if found_in_zotero else "Unreviewed",
        "Reviewer_Notes": "",
        "PubMed_URL": f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else "",
        "Last_Seen": date.today().isoformat(),
    }


def screen_article(title: str, abstract: str, section: LibrarySection) -> tuple[str, str]:
    text = f"{title} {abstract}".lower()
    title_text = title.lower()
    has_cfc_term = any(term.lower() in text for term in CFC_TERMS)
    has_strong_cfc_signal = any(term in text for term in STRONG_CFC_SIGNALS)
    has_cfc_in_title = any(term in title_text for term in STRONG_CFC_SIGNALS)
    has_cfc_gene = any(gene.lower() in text for gene in CFC_GENES)
    has_rasopathy = "rasopath" in text
    has_other_rasopathy_in_title = any(signal in title_text for signal in OTHER_RASOPATHY_SIGNALS)
    section_terms = [token.lower() for token in re.findall(r"[A-Za-z]{5,}", section.description)]
    has_section_signal = any(token in text for token in section_terms[:12])

    if has_other_rasopathy_in_title and not has_cfc_in_title:
        return "Screen out", "Title is centered on another RASopathy and does not identify CFC as the study population."
    if has_strong_cfc_signal and has_section_signal:
        return "Needs human review", "Mentions CFC and matches the section topic; confirm the article includes CFC-specific data before screening in."
    if has_strong_cfc_signal:
        return "Needs human review", "Mentions CFC, but the section-specific topic signal is weak."
    if has_rasopathy and has_cfc_gene:
        return "Needs human review", "Mentions RASopathy plus a CFC-associated gene; check whether CFC data are present."
    return "Screen out", "No clear CFC-specific signal in title or abstract."


def add_suggested_labels(df: DataFrame, model_choice: str, label_count: int) -> DataFrame:
    if df.empty:
        return df
    model = SentenceTransformer(model_choice)
    labels = list(SECTIONS)
    label_embeddings = model.encode(labels)
    texts = (df["Title"].fillna("") + ". " + df["Abstract"].fillna("")).tolist()
    article_embeddings = model.encode(texts)
    scores = util.cos_sim(article_embeddings, label_embeddings)
    suggestions = []
    for row_index, row_scores in enumerate(scores):
        primary = df.iloc[row_index]["Primary_Category"]
        pairs = sorted(
            [(float(score), label) for score, label in zip(row_scores, labels)],
            key=lambda item: item[0],
            reverse=True,
        )
        suggestions.append(", ".join([label for _, label in pairs if label != primary][:label_count]))
    df = df.copy()
    df["Suggested_Labels"] = suggestions
    return df


def add_deep_learning_review_columns(df: DataFrame, model_choice: str, label_count: int = 3) -> DataFrame:
    if df.empty:
        return df
    model = SentenceTransformer(model_choice)
    labels = list(SECTIONS)
    label_texts = [
        f"{section.name}. {section.description} Inclusion: {section.inclusion} Exclusion: {section.exclusion}"
        for section in SECTIONS.values()
    ]
    label_embeddings = model.encode(label_texts)
    texts = (df["Title"].fillna("") + ". " + df["Abstract"].fillna("")).tolist()
    article_embeddings = model.encode(texts)
    scores = util.cos_sim(article_embeddings, label_embeddings)

    top_categories = []
    top_scores = []
    suggested = []
    all_scores = []
    for row_scores in scores:
        pairs = sorted(
            [(float(score), label) for score, label in zip(row_scores, labels)],
            key=lambda item: item[0],
            reverse=True,
        )
        top_categories.append(pairs[0][1])
        top_scores.append(round(pairs[0][0], 4))
        suggested.append(", ".join(label for _, label in pairs[:label_count]))
        all_scores.append("; ".join(f"{label}:{score:.3f}" for score, label in pairs[:8]))

    output = df.copy()
    output["Deep_Learning_Top_Category"] = top_categories
    output["Deep_Learning_Top_Score"] = top_scores
    output["Deep_Learning_Suggested_Categories"] = suggested
    output["Deep_Learning_All_Category_Scores"] = all_scores
    return output


def openai_review_categories_payload() -> list[dict[str, str]]:
    return [
        {
            "category": section.name,
            "description": section.description,
            "inclusion": section.inclusion,
            "exclusion": section.exclusion,
            "include_if": CATEGORY_CLASSIFICATION_GUIDANCE.get(section.name, {}).get("include_if", []),
            "exclude_if": CATEGORY_CLASSIFICATION_GUIDANCE.get(section.name, {}).get("exclude_if", []),
            "strong_indicators": CATEGORY_CLASSIFICATION_GUIDANCE.get(section.name, {}).get("strong_indicators", []),
            "negative_examples": CATEGORY_CLASSIFICATION_GUIDANCE.get(section.name, {}).get("negative_examples", []),
        }
        for section in SECTIONS.values()
        if section.name not in EXCLUDED_COMPARISON_FOLDERS
    ]


def category_names_for_openai() -> list[str]:
    return [section.name for section in SECTIONS.values() if section.name not in EXCLUDED_COMPARISON_FOLDERS]


def normalize_openai_decision(value: object) -> str:
    text = str(value or "").strip().lower()
    if text in {"included", "include", "relevant", "screen in", "screened in"}:
        return "Included"
    if text in {"excluded", "exclude", "not relevant", "screen out", "screened out"}:
        return "Excluded"
    return "Needs review"


def normalize_openai_confidence(value: object) -> str:
    text = str(value or "").strip().lower()
    if text in {"high", "medium", "low"}:
        return text.capitalize()
    if text == "moderate":
        return "Medium"
    return "Medium"


def normalize_openai_category(value: object) -> str:
    labels = split_category_labels(value)
    for category in category_names_for_openai():
        if category in labels:
            return category
    text = str(value or "").strip()
    return text if text in category_names_for_openai() else ""


def parse_json_response(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.IGNORECASE).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start : end + 1])
        raise


def add_openai_review_columns(df: DataFrame, batch_size: int = 5, zotero_examples_text: str = "") -> DataFrame:
    if df.empty:
        return df
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is required for review comparison mode.")

    from openai import OpenAI

    client = OpenAI(api_key=api_key)
    model_name = os.getenv("OPENAI_CATEGORY_MODEL", "gpt-5.4-mini")
    output = df.copy()
    for column in [
        "OpenAI_Relevance_Decision",
        "OpenAI_Primary_Category",
        "OpenAI_Secondary_Categories",
        "OpenAI_Article_Type",
        "OpenAI_Primary_Objective",
        "OpenAI_Rationale",
        "OpenAI_Confidence",
    ]:
        output[column] = ""

    categories = openai_review_categories_payload()
    for start in range(0, len(output), batch_size):
        batch = output.iloc[start : start + batch_size]
        articles = []
        for idx, row in batch.iterrows():
            zotero_full_text = str(row.get("Zotero_Full_Text", "") or "")
            metadata_text = ". ".join(
                part
                for part in [
                    str(row.get("Title", "") or ""),
                    str(row.get("Abstract", "") or ""),
                    str(row.get("Journal", "") or ""),
                    str(row.get("DOI", "") or ""),
                ]
                if part
            )
            evidence_text = "\n\n".join(
                part
                for part in [
                    metadata_text,
                    f"ZOTERO INDEXED FULL TEXT:\n{zotero_full_text[:18000]}" if zotero_full_text else "",
                ]
                if part
            )[:20000]
            articles.append(
                {
                    "row_index": int(idx),
                    "pmid": str(row.get("PMID", "")),
                    "title": str(row.get("Title", "")),
                    "abstract": str(row.get("Abstract", ""))[:3000],
                    "year": str(row.get("Publication_Year", "")),
                    "doi": str(row.get("DOI", "")),
                    "zotero_full_text_found": str(row.get("Zotero_Full_Text_Found", "")),
                    "zotero_full_text_source": str(row.get("Zotero_Full_Text_Source", "")),
                    "article_text": evidence_text,
                    "deep_learning_top_category": str(row.get("Deep_Learning_Top_Category", "")),
                    "deep_learning_suggested_categories": str(row.get("Deep_Learning_Suggested_Categories", "")),
                    "google_sheets_found": str(row.get("Google_Sheets_Found", "")),
                    "google_sheets_decision": str(row.get("Google_Sheets_Decision", "")),
                    "zotero_found": str(row.get("Zotero_Found", "")),
                    "zotero_current_category": str(row.get("Zotero_Current_Category", "")),
                }
            )

        prompt = textwrap.dedent(
            f"""
            Classify each article for a CFC syndrome research library.

            Use the available Zotero indexed full text when present. If full text is not present,
            use title, abstract, journal, and metadata.

            Think like an experienced human curator. Do not immediately assign a category.
            Follow this hierarchical decision process for every article:

            Step 1: Determine whether the article contains substantial CFC-specific evidence.
            If no, set decision to "Excluded" and leave ai_category blank.
            If yes, continue.

            Step 2: Determine article_type using only one of:
            {json.dumps(ARTICLE_TYPES)}

            Step 3: Identify the SINGLE primary scientific objective.
            The primary objective is the hypothesis, primary research question, or main outcome measured.
            It is not every phenotype mentioned.

            Step 4: Compare only the plausible categories suggested by the article text, PubMed query matches,
            deep-learning suggestions, category guidance, and human Zotero examples.
            Do not score every category.

            Step 5: Assign exactly one primary category.

            Step 6: Assign one additional category only if a second specialty accounts for about 30% or more of the paper
            or the paper has two equally important scientific objectives. Do not add categories for phenotype mentions.

            Confidence scoring:
            - High: primary objective is obvious and one category clearly scores highest.
            - Medium: some category overlap exists, but one category is still best.
            - Low: paper could reasonably fit multiple categories or evidence is limited.
            Low-confidence non-excluded papers must use decision "Needs review".

            Output rules:
            - Exclude articles that are clearly not related to CFC, CFC-associated RAS/MAPK biology, or clinically relevant RASopathy comparison.
            - Exclude Noonan-only, Costello-only, Legius-only, NF1-only, cancer-only, somatic mutation-only, or broad pathway papers when they do not provide CFC data, a CFC-associated mutation/model, a shared RASopathy mechanism, or useful clinical comparison for CFC.
            - Inspect the supplied full text when available. If CFC is only barely mentioned, appears only in a passing disease list, appears only in a figure/table label, appears only in references, or is not substantively discussed in the aims/results/discussion, set decision to "Excluded".
            - If the article is clearly about another condition and only mentions RAS/MAPK, BRAF, MAP2K1, MAP2K2, KRAS, or RASopathy in passing, set decision to "Excluded".
            - If CFC relevance is uncertain but plausible, set decision to "Needs review" and assign the best tentative category so Lexi can check it.
            - If an article is about Noonan syndrome, Costello syndrome, Legius syndrome, NF1, or general RASopathies but may provide useful comparison, shared mechanism, Noonan-spectrum context, or CFC-relevant mutation/model information, use "Needs review" rather than excluding it.
            - Do not exclude CFC-specific clinical guidelines, diagnostic criteria papers, CFC index papers, prenatal diagnosis overviews, or CFC-focused management papers just because they are not original research. Classify these as "General and Reviews".
            - If a RASopathy-wide review or overview meaningfully discusses CFC, compares CFC with other RASopathies, or provides useful CFC clinical/genetic/management context, classify it as "General and Reviews" instead of excluding it.
            - If the article is a review, broad overview, clinical guideline, consensus statement, diagnostic summary, educational summary, or case report/case series that does not present a new mutation, new disease mechanism, new experimental result, new genotype-phenotype analysis, or previously undescribed clinical manifestation, classify it as "General and Reviews".
            - If the article is about an adult with CFC and is mainly a case report, clinical summary, or literature review, classify it as "General and Reviews" unless it presents a clearly novel specialty-specific finding.
            - If a review mainly addresses treatment options or therapeutic landscape, prefer "Treatments" as the primary category and add the organ system as an additional category when appropriate.
            - Human review examples to follow: CFC clinical management guidelines, CFC index/diagnostic criteria papers, prenatal diagnosis overviews with meaningful CFC discussion, and RASopathy reviews with substantive CFC comparison belong in "General and Reviews"; autoimmune or other RASopathy cohorts with actual CFC participants can belong in the relevant specialty; papers that only allude to CFC in one figure/table/reference should be excluded.
            - Use a specialty category only when the article contributes new original evidence or a clearly specialty-specific clinical finding for that folder.
            - Do not assign specialty folders to review articles merely because the review mentions that organ system.
            - Choose exactly one primary category. The primary category should reflect the article's main purpose, not every phenotype mentioned.
            - Secondary categories should be rare. Add at most one additional category unless the article has two clearly separate main objectives.
            - Do not add "General and Reviews" as an additional category to original research, case reports with a clear specialty finding, genotype/variant papers, or organ-system studies just because they contain background.
            - Use "General and Reviews" as primary only when the article is mainly a review, guideline, diagnostic/index paper, broad phenotype overview, or management/background paper.
            - If a paper is broad CFC clinical characterization, diagnosis, management guidance, or patient education, prefer "General and Reviews" unless there is a clearly novel specialty finding.
            - If a paper primarily identifies or validates a CFC pathogenic variant, genotype-phenotype relationship, inheritance pattern, or variant function for clinical genetics, prefer "Genetics" even if it also describes clinical features.
            - If a paper primarily studies hypertrophic cardiomyopathy, heart function, MEK inhibition for cardiac disease, arrhythmia, or cardiac surveillance, prefer "Cardiology".
            - If a paper is primarily epilepsy, EEG, seizure phenotype, or seizure treatment, prefer "Seizures" rather than Neurology.
            - If a paper is primarily broader motor milestones, behavior, cognition, motor function, hypotonia, or neuroimaging without seizure focus, prefer "Neurology" or "Development and Behavior" based on the main outcome.
            - If a paper mainly describes skin, hair, nail, eczema, keratosis, or other cutaneous findings, prefer "Dermatology" even if CFC genetics are mentioned.
            - If zotero_current_category is Exclusion or Excluded, treat that as prior human screen-out context, not as a clinical category. Judge the article independently, but do not assign Exclusion as a category.
            - If excluded, leave ai_category blank and explain the specific exclusion reason.
            - If decision is "Needs review", still provide the best tentative ai_category.
            - If included, choose one primary category in ai_category.
            - If the article clearly belongs in more than one category, mention additional categories in additional_categories.
            - Do not add multiple categories just because the article mentions multiple symptoms; only add them when the article substantially belongs in those sections.
            - Do not use Historical Articles or Conferences.

            Category priority rules:
            {json.dumps(CATEGORY_PRIORITY_RULES, indent=2)}

            Existing human Zotero categorization examples:
            {zotero_examples_text or "No Zotero examples were available for this run."}

            New inclusion/exclusion criteria and category definitions:
            {json.dumps(categories, indent=2)}

            Human-corrected examples to follow:
            - "Clinical analysis of a child with cardio-facio-cutaneous syndrome due to a de novo variant of MAP2K1 gene" -> Included, Genetics, no additional category.
            - "Cardio-facio-cutaneous syndrome: clinical features, diagnosis, and management guidelines" -> Included, General and Reviews.
            - "CFC index for the diagnosis of cardiofaciocutaneous syndrome" -> Included, General and Reviews.
            - "An Assessment of the Therapeutic Landscape for the Treatment of Heart Disease in the RASopathies" -> Included or Needs review, Treatments; Additional: Cardiology.
            - "Autoimmune disease and multiple autoantibodies in 42 patients with RASopathies" -> Included or Needs review if CFC participants are included, Allergy and Immunology.
            - "RASopathy Gene Mutations in Melanoma" -> Excluded when CFC is not substantively discussed.

            Articles:
            {json.dumps(articles, indent=2)}

            Return only valid JSON in this exact shape:
            {{
              "articles": [
                {{
                  "row_index": 0,
                  "decision": "Included",
                  "article_type": "Original research",
                  "primary_objective": "Primary objective based on title/abstract/full text.",
                  "ai_category": "Cardiology",
                  "additional_categories": ["Genetics"],
                  "confidence": "High",
                  "reasoning": "Evidence-based rationale explaining why the primary category was chosen using title/abstract/full-text evidence."
                }}
              ]
            }}

            Reasoning must explain why the primary category was chosen, not merely list topics.
            """
        ).strip()

        try:
            response = client.responses.create(model=model_name, input=prompt)
            payload = parse_json_response(getattr(response, "output_text", ""))
            assignments = payload.get("articles", [])
        except Exception as exc:
            assignments = []
            for idx, row in batch.iterrows():
                output.at[idx, "OpenAI_Relevance_Decision"] = "Possibly relevant"
                output.at[idx, "OpenAI_Primary_Category"] = str(row.get("Deep_Learning_Top_Category", ""))
                output.at[idx, "OpenAI_Secondary_Categories"] = str(row.get("Deep_Learning_Suggested_Categories", ""))
                output.at[idx, "OpenAI_Article_Type"] = ""
                output.at[idx, "OpenAI_Primary_Objective"] = ""
                output.at[idx, "OpenAI_Confidence"] = "Low"
                output.at[idx, "OpenAI_Rationale"] = f"OpenAI review failed for this batch; fallback used deep learning. Error: {type(exc).__name__}"

        for item in assignments:
            row_index = item.get("row_index")
            if row_index not in output.index:
                continue
            decision = normalize_openai_decision(item.get("decision"))
            confidence = normalize_openai_confidence(item.get("confidence"))
            if confidence == "Low" and decision != "Excluded":
                decision = "Needs review"
            if decision == "Included":
                relevance = "Relevant"
            elif decision == "Excluded":
                relevance = "Not relevant"
            else:
                relevance = "Possibly relevant"
            primary = normalize_openai_category(item.get("ai_category"))
            if relevance == "Not relevant" or primary in EXCLUDED_COMPARISON_FOLDERS or primary not in SECTIONS:
                primary = ""
            secondary = item.get("additional_categories", [])
            if isinstance(secondary, list):
                secondary_values = []
                for value in secondary:
                    category = normalize_openai_category(value)
                    if not category or category == primary:
                        continue
                    if category == "General and Reviews" and primary != "General and Reviews":
                        continue
                    if category in category_names_for_openai() and category not in secondary_values:
                        secondary_values.append(category)
                secondary_text = ", ".join(secondary_values[:1])
            else:
                secondary_text = ""
            output.at[row_index, "OpenAI_Relevance_Decision"] = relevance
            output.at[row_index, "OpenAI_Primary_Category"] = primary
            output.at[row_index, "OpenAI_Secondary_Categories"] = secondary_text
            output.at[row_index, "OpenAI_Article_Type"] = str(item.get("article_type", "") or "").strip()
            output.at[row_index, "OpenAI_Primary_Objective"] = str(item.get("primary_objective", "") or "").strip()
            output.at[row_index, "OpenAI_Confidence"] = confidence
            output.at[row_index, "OpenAI_Rationale"] = str(item.get("reasoning", "") or item.get("rationale", "") or "").strip()
    return output


def choose_system_decision(row: Any) -> tuple[str, str, str, str, str, str]:
    openai_relevance = str(row.get("OpenAI_Relevance_Decision", "") or "").strip()
    openai_category = str(row.get("OpenAI_Primary_Category", "") or "").strip()
    openai_secondary = str(row.get("OpenAI_Secondary_Categories", "") or "").strip()
    openai_confidence = str(row.get("OpenAI_Confidence", "") or "").strip()
    dl_category = str(row.get("Deep_Learning_Top_Category", "") or "").strip()
    dl_score = row.get("Deep_Learning_Top_Score", "")
    google_found = str(row.get("Google_Sheets_Found", "") or "").strip()
    zotero_found = str(row.get("Zotero_Found", "") or "").strip()
    zotero_category = str(row.get("Zotero_Current_Category", "") or "").strip()

    relevance = openai_relevance if openai_relevance in {"Relevant", "Possibly relevant", "Not relevant"} else "Possibly relevant"
    primary_category = openai_category if openai_category in SECTIONS else dl_category
    confidence = openai_confidence or "Medium"
    rationale_parts = []
    if row.get("OpenAI_Rationale"):
        rationale_parts.append(str(row.get("OpenAI_Rationale")))
    if dl_category:
        rationale_parts.append(f"Deep learning top category: {dl_category} ({dl_score}).")
    rationale = " ".join(rationale_parts).strip()

    flags = []
    if relevance == "Not relevant":
        flags.append("Not relevant")
    elif zotero_found == "No" and google_found == "Yes":
        flags.append("Previously screened but not in Zotero")
    elif zotero_found == "Yes" and zotero_category == "unfiled":
        flags.append("In Zotero but unfiled")
    elif zotero_found == "Yes":
        flags.append("Already in Zotero")
    else:
        flags.append("New relevant candidate")

    if primary_category and zotero_category and zotero_category not in {"not in Zotero", "unfiled"} and primary_category not in zotero_category:
        flags.append("Category mismatch")
    if openai_category and dl_category and openai_category != dl_category:
        flags.append("OpenAI/deep learning category mismatch")

    explanation = "; ".join(flags)
    return relevance, rationale, primary_category, openai_secondary, confidence, explanation


def finalize_review_comparison_decisions(df: DataFrame) -> DataFrame:
    output = df.copy()
    if output.empty:
        for column in [
            "System_Relevance_Decision",
            "System_Relevance_Rationale",
            "System_Primary_Category",
            "System_Secondary_Categories",
            "System_Confidence",
            "Comparison_Explanation",
            "Comparison_Flag",
            "Reviewer_Relevance",
            "Reviewer_Final_Category",
            "Reviewer_Action",
            "Reviewer_Notes",
        ]:
            output[column] = ""
        return output

    decisions = output.apply(choose_system_decision, axis=1, result_type="expand")
    decisions.columns = [
        "System_Relevance_Decision",
        "System_Relevance_Rationale",
        "System_Primary_Category",
        "System_Secondary_Categories",
        "System_Confidence",
        "Comparison_Explanation",
    ]
    for column in decisions.columns:
        output[column] = decisions[column]
    output["Comparison_Flag"] = output["Comparison_Explanation"].str.split(";").str[0]
    output["Reviewer_Relevance"] = output["System_Relevance_Decision"]
    output["Reviewer_Final_Category"] = output["System_Primary_Category"]
    output["Reviewer_Action"] = ""
    output["Reviewer_Notes"] = ""
    return output


def build_review_comparison_rows(
    records: list[dict],
    pmid_categories: dict[str, list[str]],
    zotero_index: dict[str, dict[str, dict[str, str]]],
    screening_history: dict[str, dict[str, dict[str, str]]],
) -> DataFrame:
    rows = []
    zotero_pmids = set(zotero_index.get("pmid", {}))
    for record in records:
        medline = record.get("MedlineCitation", {})
        pmid = str(medline.get("PMID", ""))
        matched_categories = pmid_categories.get(pmid, [])
        primary_section_name = matched_categories[0] if matched_categories else "General and Reviews"
        primary_section = SECTIONS.get(primary_section_name, SECTIONS["General and Reviews"])
        row = parse_pubmed_article(record, primary_section, zotero_pmids, screening_history)
        history_match = lookup_history_match(row["PMID"], row["Title"], screening_history, row.get("DOI", ""))
        zotero_match = lookup_zotero_match(row["PMID"], row["Title"], row.get("DOI", ""), zotero_index) or {}

        row["Matched_Categories"] = ", ".join(matched_categories) if matched_categories else primary_section.name
        row["Google_Sheets_Found"] = "Yes" if history_match else "No"
        row["Google_Sheets_Decision"] = history_match.get("decision", "") if history_match else ""
        row["Google_Sheets_Category"] = history_match.get("category", "") if history_match else ""
        row["Google_Sheets_Notes"] = history_match.get("notes", "") if history_match else ""
        row["Google_Sheets_Source_Row"] = history_match.get("source", "") if history_match else ""
        row["Zotero_Found"] = zotero_match.get("found", "No")
        row["Zotero_Current_Category"] = zotero_match.get("category", "not in Zotero")
        row["Zotero_File_Status"] = zotero_match.get("file_status", "not in Zotero")
        row["Zotero_Item_Key"] = zotero_match.get("item_key", "")
        rows.append(row)
    return pd.DataFrame(rows)


def build_review_deep_research_prompt(report: DataFrame) -> str:
    sample = report[report["System_Relevance_Decision"].isin(["Relevant", "Possibly relevant"])].head(75)
    article_lines = []
    for _, row in sample.iterrows():
        article_lines.append(
            f"- PMID {row.get('PMID', '')}: {row.get('Title', '')} "
            f"({row.get('Publication_Year', '')}). Relevance: {row.get('System_Relevance_Decision', '')}. "
            f"Suggested category: {row.get('System_Primary_Category', '')}. "
            f"Zotero: {row.get('Zotero_Found', '')} / {row.get('Zotero_Current_Category', '')}. "
            f"Google Sheets: {row.get('Google_Sheets_Found', '')} / {row.get('Google_Sheets_Decision', '')}. "
            f"URL: {row.get('PubMed_URL', '')}"
        )
    articles = "\n".join(article_lines) if article_lines else "No relevant or possibly relevant candidates were found."
    return textwrap.dedent(
        f"""
        Conduct a deep research quality check for a CFC syndrome review-comparison workbook covering PubMed articles from {REVIEW_COMPARISON_YEARS[0]}-{REVIEW_COMPARISON_YEARS[1]}.

        The workbook compares OpenAI category assignment, deep learning similarity, Google Sheets screening history, and Zotero folder status.

        Category definitions and inclusion/exclusion criteria:
        {json.dumps(openai_review_categories_payload(), indent=2)}

        Candidate articles:
        {articles}

        Produce:
        1. A concise assessment of whether the relevance and category assignments look reasonable.
        2. Articles most likely missing from Zotero, especially those found in Google Sheets but not Zotero.
        3. Articles whose proposed category conflicts with Zotero or Google Sheets.
        4. Any categories whose criteria may need refinement.
        Cite PubMed IDs for article-specific claims.
        """
    ).strip()


def review_comparison_export_frame(report: DataFrame) -> DataFrame:
    export = report.copy()
    blank = pd.Series([""] * len(export), index=export.index)
    google_decision = export["Google_Sheets_Decision"] if "Google_Sheets_Decision" in export else blank
    google_screen = google_decision.fillna("").map(screening_direction)
    zotero_found = export["Zotero_Found"].fillna("") if "Zotero_Found" in export else blank
    zotero_category = export["Zotero_Current_Category"].fillna("") if "Zotero_Current_Category" in export else blank
    zotero_screen = zotero_category.map(
        lambda value: (
            "out"
            if "Excluded" in split_category_labels(value)
            else (
                "in"
                if str(value or "").strip()
                and str(value or "").strip() not in {"not in Zotero", "unfiled"}
                and bool(split_category_labels(value))
                else ""
            )
        )
    ).where(zotero_found.eq("Yes"), "")
    human_screen = google_screen.where(google_screen.ne(""), zotero_screen)
    if "OpenAI_Relevance_Decision" in export:
        openai_relevance = export["OpenAI_Relevance_Decision"].fillna("")
    elif "System_Relevance_Decision" in export:
        openai_relevance = export["System_Relevance_Decision"].fillna("")
    else:
        openai_relevance = blank
    openai_screen = openai_relevance.map(screening_direction)
    export["Human_OpenAI_Match"] = [
        "Yes" if human and ai and human == ai else ("No" if human and ai else "")
        for human, ai in zip(human_screen, openai_screen)
    ]
    export["OpenAI_Screening_Display"] = openai_relevance.map(
        {
            "Relevant": "Screen in",
            "Possibly relevant": "Needs review",
            "Not relevant": "Screen out",
        }
    ).fillna(openai_relevance)
    export["Zotero_Category_Display"] = zotero_category.where(zotero_found.eq("Yes"), "")
    openai_primary = export["OpenAI_Primary_Category"].fillna("") if "OpenAI_Primary_Category" in export else blank
    openai_secondary = export["OpenAI_Secondary_Categories"].fillna("") if "OpenAI_Secondary_Categories" in export else blank
    export["OpenAI_Assigned_Category_Display"] = [
        "" if str(relevance).strip() == "Not relevant" else comparison_category_text(primary, secondary)
        for relevance, primary, secondary in zip(openai_relevance, openai_primary, openai_secondary)
    ]
    for source_column, _ in REVIEW_COMPARISON_EXPORT_COLUMNS:
        if source_column not in export:
            export[source_column] = ""
    source_columns = [source_column for source_column, _ in REVIEW_COMPARISON_EXPORT_COLUMNS]
    header_map = dict(REVIEW_COMPARISON_EXPORT_COLUMNS)
    return export[source_columns].rename(columns=header_map)


def write_review_comparison_workbook(
    path: Path,
    report: DataFrame,
    run_log: DataFrame,
    deep_research_prompt: str,
    openai_run: dict[str, str] | None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    report = report.copy()
    sort_columns = [column for column in ("System_Primary_Category", "System_Relevance_Decision", "Title") if column in report]
    if sort_columns:
        report = report.sort_values(sort_columns, kind="stable").reset_index(drop=True)

    relevant = report[report["System_Relevance_Decision"].isin(["Relevant", "Possibly relevant"])].copy()
    not_relevant = report[report["System_Relevance_Decision"].eq("Not relevant")].copy()
    needs_review = report[
        report["System_Relevance_Decision"].eq("Possibly relevant")
        | report["Comparison_Explanation"].str.contains("mismatch|unfiled|Previously screened", case=False, na=False)
    ].copy()
    summary = (
        report.groupby(["System_Primary_Category", "System_Relevance_Decision"], dropna=False)
        .size()
        .reset_index(name="Article_Count")
        .sort_values(["System_Primary_Category", "System_Relevance_Decision"])
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        review_comparison_export_frame(report).to_excel(writer, sheet_name="Review_Comparison", index=False)
        review_comparison_export_frame(relevant).to_excel(writer, sheet_name="Relevant_Candidates", index=False)
        review_comparison_export_frame(not_relevant).to_excel(writer, sheet_name="Not_Relevant", index=False)
        review_comparison_export_frame(needs_review).to_excel(writer, sheet_name="Needs_Review", index=False)
        summary.rename(
            columns={
                "System_Primary_Category": "Suggested Category",
                "System_Relevance_Decision": "Relevance",
            }
        ).to_excel(writer, sheet_name="Category_Summary", index=False)
        criteria_frame().to_excel(writer, sheet_name="Criteria", index=False)
        pd.DataFrame({"Deep_Research_Brief": [deep_research_prompt]}).to_excel(writer, sheet_name="Deep_Research_Brief", index=False)
        if openai_run:
            pd.DataFrame([openai_run]).to_excel(writer, sheet_name="OpenAI_Run", index=False)
        run_log.to_excel(writer, sheet_name="Run_Log", index=False)
    apply_review_comparison_formatting(path)


def apply_review_comparison_formatting(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = load_workbook(path)
    sheets_to_format = ["Review_Comparison", "Relevant_Candidates", "Not_Relevant", "Needs_Review"]
    relevance_values = "Relevant,Possibly relevant,Not relevant"
    action_values = "Needs reviewed,Approved,Not approved"

    relevance_fills = {
        "Relevant": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "Possibly relevant": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Not relevant": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    }
    for sheet_name in sheets_to_format:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        max_row = max(ws.max_row, 2)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)

        if "Reviewer Relevance" in headers:
            col = get_column_letter(headers.index("Reviewer Relevance") + 1)
            validation = DataValidation(type="list", formula1=f'"{relevance_values}"', allow_blank=True)
            ws.add_data_validation(validation)
            validation.add(f"{col}2:{col}{max_row}")
        if "Reviewer Action" in headers:
            col = get_column_letter(headers.index("Reviewer Action") + 1)
            validation = DataValidation(type="list", formula1=f'"{action_values}"', allow_blank=True)
            ws.add_data_validation(validation)
            validation.add(f"{col}2:{col}{max_row}")
        if "Relevance" in headers:
            decision_col = get_column_letter(headers.index("Relevance") + 1)
            decision_range = f"{decision_col}2:{decision_col}{max_row}"
            for decision, fill in relevance_fills.items():
                ws.conditional_formatting.add(
                    decision_range,
                    FormulaRule(formula=[f'${decision_col}2="{decision}"'], fill=fill),
                )
        if "Match?" in headers:
            match_col = get_column_letter(headers.index("Match?") + 1)
            match_range = f"{match_col}2:{match_col}{max_row}"
            ws.conditional_formatting.add(
                match_range,
                FormulaRule(formula=[f'${match_col}2="Yes"'], fill=PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")),
            )
        for category_header in ("If in Zotero - Category", "OpenAI Assigned Category"):
            if category_header in headers:
                category_col = headers.index(category_header) + 1
                for row in range(2, max_row + 1):
                    category = str(ws.cell(row=row, column=category_col).value or "")
                    color = CATEGORY_COLORS.get(category, "FFFFFF")
                    ws.cell(row=row, column=category_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        widths = {
            "PubMed ID": 12,
            "Article Title": 60,
            "Date": 14,
            "Relevance": 18,
            "Found in Sheets?": 18,
            "Found in Zotero?": 18,
            "Match?": 12,
            "OpenAI Screening In or Out": 26,
            "If in Zotero - Category": 30,
            "OpenAI Assigned Category": 26,
        }
        for index, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(index)].width = widths.get(str(header), 18)

    if "Category_Summary" in wb.sheetnames:
        ws = wb["Category_Summary"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
    wb.save(path)


def run_review_comparison(args: argparse.Namespace) -> None:
    from_year, to_year = REVIEW_COMPARISON_YEARS
    Entrez.email = require_env("ENTREZ_EMAIL")
    Entrez.api_key = os.getenv("NCBI_API_KEY") or None
    require_env("OPENAI_API_KEY")

    screening_history = load_screening_history(args.screening_history)
    zotero_index = {}
    zotero_examples_text = ""
    if not args.skip_zotero:
        zotero_group_id = require_env("ZOTERO_GROUP_ID")
        zotero_api_key = require_env("ZOTERO_API_KEY")
        zotero_index = fetch_zotero_index(zotero_group_id, zotero_api_key)
        zotero_examples_text = build_zotero_assignment_examples(zotero_group_id, zotero_api_key)

    pmid_categories: dict[str, list[str]] = {}
    section_counts: list[dict[str, object]] = []
    for section in SECTIONS.values():
        section_pmids = search_pubmed(section.query, args.retmax, from_year, to_year)
        time.sleep(0.4)
        section_counts.append({"Section": section.name, "PubMed_Records_Found": len(section_pmids)})
        for pmid in section_pmids:
            pmid_categories.setdefault(pmid, [])
            if section.name not in pmid_categories[pmid]:
                pmid_categories[pmid].append(section.name)

    pubmed_pmids = list(pmid_categories)
    records = fetch_pubmed_records(pubmed_pmids) if pubmed_pmids else []
    report = build_review_comparison_rows(records, pmid_categories, zotero_index, screening_history)
    report = filter_report_year_range(report, from_year, to_year)
    if not args.skip_zotero:
        report = add_zotero_fulltext_columns(
            report,
            zotero_group_id,
            zotero_api_key,
        )
    report = add_deep_learning_review_columns(report, args.embedding_model, max(args.suggested_labels, 3))
    report = add_openai_review_columns(report, args.openai_review_batch_size, zotero_examples_text)
    report = finalize_review_comparison_decisions(report)

    deep_research_prompt = build_review_deep_research_prompt(report)
    openai_run = None if args.skip_openai_deep_research else launch_openai_deep_research(deep_research_prompt)
    log_rows = [
        {"Field": "Mode", "Value": "2017-2022 review comparison"},
        {"Field": "Publication date filter", "Value": format_year_filter(from_year, to_year)},
        {"Field": "OpenAI category model", "Value": os.getenv("OPENAI_CATEGORY_MODEL", "gpt-5.4-mini")},
        {"Field": "Unique PubMed records found", "Value": len(pubmed_pmids)},
        {"Field": "Rows in report", "Value": len(report)},
        {"Field": "Google Sheets matches", "Value": int((report.get("Google_Sheets_Found") == "Yes").sum()) if not report.empty else 0},
        {"Field": "Zotero matches", "Value": int((report.get("Zotero_Found") == "Yes").sum()) if not report.empty else 0},
        {"Field": "Zotero full text matches", "Value": int((report.get("Zotero_Full_Text_Found") == "Yes").sum()) if "Zotero_Full_Text_Found" in report else 0},
        {"Field": "Zotero categorization examples used", "Value": "Yes" if zotero_examples_text else "No"},
        {"Field": "Generated", "Value": time.ctime()},
    ]
    for item in section_counts:
        log_rows.append({"Field": f"PubMed records found: {item['Section']}", "Value": item["PubMed_Records_Found"]})
    run_log = pd.DataFrame(
        log_rows
    )
    output_path = Path(args.output)
    write_review_comparison_workbook(output_path, report, run_log, deep_research_prompt, openai_run)

    print("Mode: 2017-2022 review comparison")
    print(f"Publication date filter: {format_year_filter(from_year, to_year)}")
    print(f"Unique PubMed records found: {len(pubmed_pmids)}")
    print(f"Rows written: {len(report)}")
    print(f"Google Sheets matches: {int((report.get('Google_Sheets_Found') == 'Yes').sum()) if not report.empty else 0}")
    print(f"Zotero matches: {int((report.get('Zotero_Found') == 'Yes').sum()) if not report.empty else 0}")
    if "Zotero_Full_Text_Found" in report:
        print(f"Zotero full text matches: {int((report.get('Zotero_Full_Text_Found') == 'Yes').sum())}")
    if openai_run:
        print(f"OpenAI deep research submitted: {openai_run.get('response_id', '')} ({openai_run.get('status', '')})")
    print(f"Workbook written: {output_path.resolve()}")


def merge_with_existing(existing: DataFrame, new_rows: DataFrame) -> DataFrame:
    if existing.empty:
        return new_rows
    for column in FINAL_COLUMNS:
        if column not in existing:
            existing[column] = ""
    existing = existing[FINAL_COLUMNS].copy()
    existing_pmids = set(existing["PMID"].astype(str))
    truly_new = new_rows[~new_rows["PMID"].astype(str).isin(existing_pmids)]
    refreshed = pd.concat([existing, truly_new], ignore_index=True)
    return refreshed.drop_duplicates(subset=["PMID"], keep="first")


def filter_report_year_range(report: DataFrame, from_year: int | None, to_year: int | None) -> DataFrame:
    if (not from_year and not to_year) or report.empty or "Publication_Year" not in report:
        return report
    filtered = report.copy()
    years = pd.to_numeric(filtered["Publication_Year"], errors="coerce")
    mask = pd.Series(True, index=filtered.index)
    if from_year:
        mask = mask & (years >= from_year)
    if to_year:
        mask = mask & (years <= to_year)
    return filtered[mask].reset_index(drop=True)


def load_existing(path: Path) -> tuple[DataFrame, int]:
    if not path.exists():
        return pd.DataFrame(columns=FINAL_COLUMNS), 1
    existing = pd.read_excel(path, sheet_name="Review_Report", dtype={"PMID": str})
    run_count = 1
    try:
        instructions = pd.read_excel(path, sheet_name="Instructions", header=None)
        line = instructions[instructions[0].astype(str).str.contains("Run Count:", na=False)].iloc[0, 0]
        run_count = int(re.search(r"\d+", str(line)).group()) + 1
    except Exception:
        run_count = 1
    return existing, run_count


def criteria_frame() -> DataFrame:
    return pd.DataFrame(
        [
            {
                "Section": section.name,
                "Description": section.description,
                "Inclusion": section.inclusion,
                "Exclusion": section.exclusion,
                "PubMed_Query": section.query,
            }
            for section in SECTIONS.values()
        ]
    )


def build_deep_research_prompt(section: LibrarySection | None, recent_df: DataFrame) -> str:
    article_lines = []
    for _, row in recent_df.head(50).iterrows():
        article_lines.append(
            f"- PMID {row['PMID']}: {row['Title']} ({row.get('Publication_Year', '')}). "
            f"Decision: {row.get('Eligibility_Decision', '')}. URL: {row.get('PubMed_URL', '')}"
        )
    articles = "\n".join(article_lines) if article_lines else "No new candidate articles were found in this run."
    if section:
        scope = f"the Zotero section: {section.name}"
        section_context = textwrap.dedent(
            f"""
            Section description:
            {section.description}

            Inclusion criteria:
            {section.inclusion}

            Exclusion criteria:
            {section.exclusion}
            """
        ).strip()
    else:
        scope = "all Zotero library sections in the CFC research library"
        section_context = (
            "Use each article's Primary_Category and Matched_Categories fields to keep the original "
            "Zotero folder structure intact while reviewing the combined update."
        )

    return textwrap.dedent(
        f"""
        Conduct a deep research review update for {scope}.

        {section_context}

        General inclusion criteria:
        {json.dumps(GENERAL_INCLUSION_CRITERIA, indent=2)}

        General exclusion criteria:
        {json.dumps(GENERAL_EXCLUSION_CRITERIA, indent=2)}

        Candidate articles from the latest PubMed/Zotero update:
        {articles}

        Produce:
        1. A concise summary of the most important new findings.
        2. A table of articles that should be added to Zotero, with rationale.
        3. Articles that should be excluded or require human review, with rationale.
        4. Any missing search terms or adjacent concepts that should be considered for the next update.
        Use PubMed IDs and source links for every article-specific claim.
        """
    ).strip()


def launch_openai_deep_research(prompt: str) -> dict[str, str]:
    """Start an OpenAI deep research run and return lightweight metadata."""
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is required because deep research runs by default.")

    from openai import OpenAI

    client = OpenAI(api_key=api_key)
    response = client.responses.create(
        model=os.getenv("OPENAI_DEEP_RESEARCH_MODEL", "o3-deep-research"),
        input=prompt,
        background=True,
        tools=[
            {"type": "web_search_preview"},
            {"type": "code_interpreter", "container": {"type": "auto"}},
        ],
    )
    return {
        "response_id": getattr(response, "id", ""),
        "model": getattr(response, "model", os.getenv("OPENAI_DEEP_RESEARCH_MODEL", "o3-deep-research")),
        "status": getattr(response, "status", "submitted"),
        "submitted_at": time.ctime(),
    }


def write_workbook(
    path: Path,
    report: DataFrame,
    run_count: int,
    section: LibrarySection | None,
    prompt: str,
    from_year: int | None,
    to_year: int | None,
    openai_run: dict[str, str] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    report = report.copy()
    if "Reviewer_Decision" in report:
        report["Reviewer_Decision"] = report["Reviewer_Decision"].replace("", pd.NA).fillna("Needs reviewed")
    sort_columns = [column for column in ("Primary_Category", "Title", "Publication_Year") if column in report]
    if sort_columns:
        report = report.sort_values(sort_columns, kind="stable").reset_index(drop=True)
    instructions = pd.DataFrame(
        [
            [f"Report generated on: {time.ctime()}"],
            [f"Run Count: {run_count}"],
            [f"Primary Category: {section.name if section else 'All categories'}"],
            [f"Publication Date Filter: {format_year_filter(from_year, to_year)}"],
            [""],
            ["How to Use This File"],
            ["1. Review_Report contains candidate articles and preserves Reviewer_Notes across runs."],
            ["2. Criteria contains the library descriptions plus inclusion/exclusion criteria for every Zotero section."],
            ["3. Deep_Research_Brief contains a prompt you can run with OpenAI deep research or an agent workflow."],
            ["4. Use Reviewer_Decision to mark Needs reviewed, Approved, or Not approved."],
            ["5. Re-run this program periodically; only unseen PubMed records are appended."],
        ]
    )
    run_log = pd.DataFrame(
        [
            {"Field": "Category", "Value": section.name if section else "All categories"},
            {"Field": "PubMed Query", "Value": section.query if section else "Multiple category queries"},
            {"Field": "Publication date filter", "Value": format_year_filter(from_year, to_year)},
            {"Field": "Rows in report", "Value": len(report)},
            {"Field": "Generated", "Value": time.ctime()},
        ]
    )
    category_summary = (
        report.groupby("Primary_Category", dropna=False)
        .size()
        .reset_index(name="Article_Count")
        .sort_values(["Primary_Category"])
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        report[FINAL_COLUMNS].to_excel(writer, sheet_name="Review_Report", index=False)
        category_summary.to_excel(writer, sheet_name="Category_Summary", index=False)
        criteria_frame().to_excel(writer, sheet_name="Criteria", index=False)
        pd.DataFrame({"Deep_Research_Brief": [prompt]}).to_excel(writer, sheet_name="Deep_Research_Brief", index=False)
        if openai_run:
            pd.DataFrame([openai_run]).to_excel(writer, sheet_name="OpenAI_Run", index=False)
        run_log.to_excel(writer, sheet_name="Run_Log", index=False)
        instructions.to_excel(writer, sheet_name="Instructions", index=False, header=False)
    apply_review_dropdown_formatting(path)


def apply_review_dropdown_formatting(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = load_workbook(path)
    ws = wb["Review_Report"]
    headers = [cell.value for cell in ws[1]]
    if "Reviewer_Decision" not in headers:
        wb.save(path)
        return

    decision_col = headers.index("Reviewer_Decision") + 1
    decision_letter = get_column_letter(decision_col)
    primary_category_col = headers.index("Primary_Category") + 1 if "Primary_Category" in headers else None
    matched_category_col = headers.index("Matched_Categories") + 1 if "Matched_Categories" in headers else None
    max_row = max(ws.max_row, 2)
    decision_range = f"{decision_letter}2:{decision_letter}{max_row}"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(REVIEWER_DECISIONS)}"',
        allow_blank=True,
    )
    ws.add_data_validation(validation)
    validation.add(decision_range)

    fills = {
        "Needs reviewed": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Approved": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        "Not approved": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
    }
    for decision, fill in fills.items():
        ws.conditional_formatting.add(
            decision_range,
            FormulaRule(formula=[f'${decision_letter}2="{decision}"'], fill=fill),
        )

    if primary_category_col:
        for row in range(2, max_row + 1):
            category_cell = ws.cell(row=row, column=primary_category_col)
            color = CATEGORY_COLORS.get(str(category_cell.value), "FFFFFF")
            category_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            category_cell.font = Font(bold=True)
            if matched_category_col:
                ws.cell(row=row, column=matched_category_col).fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid",
                )

    summary_ws = wb["Category_Summary"] if "Category_Summary" in wb.sheetnames else None
    if summary_ws:
        summary_headers = [cell.value for cell in summary_ws[1]]
        if "Primary_Category" in summary_headers:
            summary_category_col = summary_headers.index("Primary_Category") + 1
            for row in range(2, summary_ws.max_row + 1):
                category_cell = summary_ws.cell(row=row, column=summary_category_col)
                color = CATEGORY_COLORS.get(str(category_cell.value), "FFFFFF")
                category_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                category_cell.font = Font(bold=True)
        summary_ws.freeze_panes = "A2"
        summary_ws.auto_filter.ref = summary_ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)

    widths = {
        "PMID": 12,
        "Title": 55,
        "Primary_Category": 24,
        "Matched_Categories": 38,
        "Eligibility_Decision": 20,
        "Reviewer_Decision": 18,
        "Reviewer_Notes": 36,
        "PubMed_URL": 34,
    }
    for index, header in enumerate(headers, start=1):
        width = widths.get(str(header), 18)
        ws.column_dimensions[get_column_letter(index)].width = width

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update a CFC syndrome literature review report.")
    parser.add_argument("--category", default="Dermatology", help="Zotero/library section to update.")
    parser.add_argument(
        "--all-categories",
        action="store_true",
        help="Search every configured Zotero/library section and export one combined workbook.",
    )
    parser.add_argument(
        "--review-comparison-2017-2022",
        action="store_true",
        help="Create the 2017-2022 all-category comparison workbook using PubMed, OpenAI, deep learning, Google Sheets history, and Zotero status.",
    )
    parser.add_argument("--output", default="reports/CFC_Master_Review_Report.xlsx", help="Workbook output path.")
    parser.add_argument("--retmax", type=int, default=10000, help="Maximum PubMed IDs to return.")
    parser.add_argument(
        "--since-year",
        type=int,
        default=2025,
        help="Only search PubMed records published from this year onward. Use 0 to search all years.",
    )
    parser.add_argument(
        "--from-year",
        type=int,
        help="First publication year to include. Overrides --since-year when provided.",
    )
    parser.add_argument(
        "--to-year",
        type=int,
        help="Last publication year to include. Use with --from-year for a bounded date range.",
    )
    parser.add_argument(
        "--embedding-model",
        default="microsoft/BiomedNLP-PubMedBERT-base-uncased-abstract-fulltext",
        help="Sentence-transformers model used for suggested labels.",
    )
    parser.add_argument("--suggested-labels", type=int, default=2, help="Number of secondary labels to suggest.")
    parser.add_argument(
        "--openai-review-batch-size",
        type=int,
        default=5,
        help="Number of articles to send per OpenAI category-review request in review comparison mode.",
    )
    parser.add_argument("--skip-zotero", action="store_true", help="Skip Zotero API lookup and mark all as not found in Zotero.")
    parser.add_argument(
        "--screening-history",
        default=DEFAULT_SCREENING_HISTORY_URL,
        help="Previously screened .xlsx/.csv/.tsv file or accessible Google Sheets URL. Matches by PMID first, then title.",
    )
    parser.add_argument(
        "--include-previously-screened",
        action="store_true",
        help="Keep screening-history matches in the new report instead of filtering them out.",
    )
    parser.add_argument(
        "--skip-openai-deep-research",
        action="store_true",
        help="Do not submit the generated brief to OpenAI deep research. Use only for testing.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    load_env_file()
    load_runtime_dependencies()
    if args.review_comparison_2017_2022:
        if args.output == "reports/CFC_Master_Review_Report.xlsx":
            args.output = "reports/CFC_2017_2022_Review_Comparison_Human_Aligned.xlsx"
        run_review_comparison(args)
        return

    category = normalize_category(args.category)
    run_all_categories = args.all_categories or category.lower() in {"all", "all categories"}
    from_year = args.from_year if args.from_year is not None else (args.since_year or None)
    to_year = args.to_year
    if not run_all_categories and category not in SECTIONS:
        choices = ", ".join(sorted(SECTIONS))
        raise SystemExit(f"Unknown category '{args.category}'. Choose one of: {choices}")

    section = None if run_all_categories else SECTIONS[category]
    sections_to_run = list(SECTIONS.values()) if run_all_categories else [section]
    Entrez.email = require_env("ENTREZ_EMAIL")
    Entrez.api_key = os.getenv("NCBI_API_KEY") or None
    if not args.skip_openai_deep_research:
        require_env("OPENAI_API_KEY")
    output_path = Path(args.output)

    existing, run_count = load_existing(output_path)
    existing_pmids = set(existing.get("PMID", pd.Series(dtype=str)).astype(str))
    screening_history = load_screening_history(args.screening_history)
    history_pmids = set(screening_history["pmid"])

    zotero_pmids = set()
    if not args.skip_zotero:
        zotero_pmids = fetch_zotero_pmids(require_env("ZOTERO_GROUP_ID"), require_env("ZOTERO_API_KEY"))

    pmid_categories: dict[str, list[str]] = {}
    section_counts: list[dict[str, object]] = []
    for active_section in sections_to_run:
        section_pmids = search_pubmed(active_section.query, args.retmax, from_year, to_year)
        section_counts.append({"Section": active_section.name, "PubMed_Records_Found": len(section_pmids)})
        for pmid in section_pmids:
            pmid_categories.setdefault(pmid, [])
            if active_section.name not in pmid_categories[pmid]:
                pmid_categories[pmid].append(active_section.name)

    pubmed_pmids = list(pmid_categories)
    new_pmids = [
        pmid
        for pmid in pubmed_pmids
        if pmid not in existing_pmids and (args.include_previously_screened or pmid not in history_pmids)
    ]
    records = fetch_pubmed_records(new_pmids) if new_pmids else []
    rows = []
    for record in records:
        pmid = str(record.get("MedlineCitation", {}).get("PMID", ""))
        matched_categories = pmid_categories.get(pmid, [])
        primary_section_name = matched_categories[0] if matched_categories else sections_to_run[0].name
        primary_section = SECTIONS[primary_section_name]
        row = parse_pubmed_article(record, primary_section, zotero_pmids, screening_history)
        row["Matched_Categories"] = ", ".join(matched_categories) if matched_categories else primary_section.name
        rows.append(row)
    if not args.include_previously_screened:
        rows = [row for row in rows if not row["History_Match"]]
    new_df = pd.DataFrame(rows, columns=FINAL_COLUMNS)
    new_df = filter_report_year_range(new_df, from_year, to_year)
    new_df = add_suggested_labels(new_df, args.embedding_model, args.suggested_labels)
    report = merge_with_existing(existing, new_df)
    report = filter_report_year_range(report, from_year, to_year)
    prompt = build_deep_research_prompt(section, new_df)
    openai_run = None if args.skip_openai_deep_research else launch_openai_deep_research(prompt)
    write_workbook(output_path, report, run_count, section, prompt, from_year, to_year, openai_run)

    print(f"Category: {section.name if section else 'All categories'}")
    print(f"Publication date filter: {format_year_filter(from_year, to_year)}")
    print(f"Unique PubMed records found: {len(pubmed_pmids)}")
    if run_all_categories:
        print("Section counts:")
        for item in section_counts:
            print(f"  - {item['Section']}: {item['PubMed_Records_Found']}")
    if args.screening_history:
        print(f"Previously screened PMIDs loaded: {len(history_pmids)}")
    if openai_run:
        print(f"OpenAI deep research submitted: {openai_run.get('response_id', '')} ({openai_run.get('status', '')})")
    print(f"New records appended: {len(new_df)}")
    print(f"Workbook written: {output_path.resolve()}")


if __name__ == "__main__":
    main()
