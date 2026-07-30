"""Calculate reviewed-only multilabel performance metrics for CFC article classification.

Default input:
    reports/CFC_2017_2022_Review_Comparison.xlsx

Default output:
    reports/CFC_2017_2022_2x2_Matrix.xlsx
"""

from __future__ import annotations

import argparse
import re
import tempfile
import time
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import urlretrieve

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = "reports/CFC_2017_2022_Review_Comparison.xlsx"
DEFAULT_OUTPUT = "reports/CFC_2017_2022_2x2_Matrix.xlsx"
DEFAULT_SHEET = "Review_Comparison"
DEFAULT_SCREENING_HISTORY_URL = (
    "https://docs.google.com/spreadsheets/d/1BUvWcV6XgYiOL3cCrYAHjkccb24C38OK/"
    "edit?usp=sharing&ouid=106518116377917721454&rtpof=true&sd=true"
)

CATEGORIES = [
    "Allergy and Immunology",
    "Cardiology",
    "Dermatology",
    "Development and Behavior",
    "Endocrinology",
    "Gastroenterology",
    "General and Reviews",
    "Genetics",
    "Growth",
    "Gynecology",
    "Neurology",
    "Oncology",
    "Ophthalmology",
    "Orthopedic",
    "Otolaryngology",
    "Pulmonology",
    "Research Studies",
    "Seizures",
    "Treatments",
]

CATEGORY_COLORS = {
    "Allergy and Immunology": "D9EAD3",
    "Cardiology": "F4CCCC",
    "Dermatology": "FCE5CD",
    "Development and Behavior": "D9D2E9",
    "Endocrinology": "D0E0E3",
    "Gastroenterology": "FFF2CC",
    "General and Reviews": "EADCF8",
    "Genetics": "CFE2F3",
    "Growth": "E2F0CB",
    "Gynecology": "FCE4EC",
    "Neurology": "C9DAF8",
    "Oncology": "E6B8AF",
    "Ophthalmology": "D9EAD3",
    "Orthopedic": "EAD7C2",
    "Otolaryngology": "D0E0E3",
    "Pulmonology": "D9EAF7",
    "Research Studies": "E6E6E6",
    "Seizures": "D9D2E9",
    "Treatments": "D5A6BD",
}

CATEGORY_ALIASES = {
    "allergy": "Allergy and Immunology",
    "allergy and immunology": "Allergy and Immunology",
    "cardio": "Cardiology",
    "cardiology": "Cardiology",
    "derm": "Dermatology",
    "dermatology": "Dermatology",
    "development": "Development and Behavior",
    "development and behavior": "Development and Behavior",
    "behavior": "Development and Behavior",
    "endocrinology": "Endocrinology",
    "gastro": "Gastroenterology",
    "gastroenterology": "Gastroenterology",
    "general": "General and Reviews",
    "general_and_reviews": "General and Reviews",
    "general and reviews": "General and Reviews",
    "reviews": "General and Reviews",
    "genetic": "Genetics",
    "genetics": "Genetics",
    "growth": "Growth",
    "gynecology": "Gynecology",
    "neuro": "Neurology",
    "neurology": "Neurology",
    "oncology": "Oncology",
    "ophthalmology": "Ophthalmology",
    "orthopedic": "Orthopedic",
    "orthopedics": "Orthopedic",
    "ent": "Otolaryngology",
    "otolaryngology ent": "Otolaryngology",
    "otolaryngology_ent": "Otolaryngology",
    "otolaryngology": "Otolaryngology",
    "pulmonology": "Pulmonology",
    "research": "Research Studies",
    "research studies": "Research Studies",
    "seizure": "Seizures",
    "seizures": "Seizures",
    "treatment": "Treatments",
    "treatments": "Treatments",
    "exclude": "Excluded",
    "excluded": "Excluded",
    "exclusion": "Excluded",
}


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text.lower())
    return re.sub(r"\s+", " ", text).strip()


def normalize_pmid(value: object) -> str:
    if value is None:
        return ""
    match = re.search(r"\b(\d{4,})\b", str(value))
    return match.group(1) if match else ""


def normalize_doi(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("_", " ")
    text = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", text)
    text = re.sub(r"^doi:\s*", "", text)
    return text.strip().rstrip(".")


def find_column(df: pd.DataFrame, candidates: tuple[str, ...]) -> str | None:
    normalized = {normalize_text(column): column for column in df.columns}
    for candidate in candidates:
        key = normalize_text(candidate)
        if key in normalized:
            return normalized[key]
    return None


def materialize_source(source: str) -> Path:
    parsed = urlparse(source)
    if parsed.scheme in {"http", "https"}:
        match = re.search(r"/spreadsheets/d/([^/]+)", source)
        download_url = source
        suffix = ".xlsx"
        if match:
            download_url = f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"
        target = Path(tempfile.gettempdir()) / f"cfc_human_screening_{int(time.time())}{suffix}"
        urlretrieve(download_url, target)
        return target
    return Path(source)


def screening_direction(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    out_terms = ("screen out", "screened out", "exclude", "excluded", "not approved", "not relevant", "no")
    in_terms = ("screen in", "screened in", "include", "included", "approved", "relevant", "yes")
    if any(term in text for term in out_terms):
        return "out"
    if any(term in text for term in in_terms):
        return "in"
    return ""


def load_human_history(source: str | None) -> dict[str, dict[str, dict[str, str]]]:
    empty = {"pmid": {}, "doi": {}, "title": {}}
    if not source:
        return empty

    path = materialize_source(source)
    if not path.exists():
        raise RuntimeError(f"Human screening file was not found: {path}")

    if path.suffix.lower() in {".xlsx", ".xls"}:
        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
        frames = [frame.assign(_history_sheet=name) for name, frame in sheets.items()]
        history_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    elif path.suffix.lower() == ".csv":
        history_df = pd.read_csv(path, dtype=str)
        history_df["_history_sheet"] = path.name
    elif path.suffix.lower() == ".tsv":
        history_df = pd.read_csv(path, sep="\t", dtype=str)
        history_df["_history_sheet"] = path.name
    else:
        raise RuntimeError("Human screening history must be .xlsx, .xls, .csv, .tsv, or an accessible Google Sheets URL.")

    if history_df.empty:
        return empty

    pmid_col = find_column(history_df, ("PMID", "PubMed ID", "PubMed_ID", "PMID Number"))
    doi_col = find_column(history_df, ("DOI", "Digital Object Identifier"))
    title_col = find_column(history_df, ("Title", "Article Title", "Article_Title"))
    decision_col = find_column(history_df, ("Eligibility_Decision", "Eligibility Decision", "Decision", "Review_Status", "Review Status"))
    category_col = find_column(history_df, ("Category", "Primary_Category", "Primary Category", "Folder", "Zotero Folder", "Suggested_Labels", "Suggested Labels"))
    notes_col = find_column(history_df, ("Notes", "Reviewer_Notes", "Reviewer Notes", "Rationale", "Comment", "Comments"))

    history = {"pmid": {}, "doi": {}, "title": {}}
    for idx, row in history_df.iterrows():
        pmid = normalize_pmid(row.get(pmid_col)) if pmid_col else ""
        doi = normalize_doi(row.get(doi_col)) if doi_col else ""
        title = normalize_text(row.get(title_col)) if title_col else ""
        decision = str(row.get(decision_col, "") or "").strip() if decision_col else ""
        category = str(row.get(category_col, "") or "").strip() if category_col else ""
        notes = str(row.get(notes_col, "") or "").strip() if notes_col else ""
        record = {
            "decision": decision,
            "screen": screening_direction(decision),
            "category": category,
            "notes": notes,
            "source": f"{path.name}:{row.get('_history_sheet', 'Sheet')}:{idx + 2}",
        }
        if pmid:
            history["pmid"][pmid] = record
        if doi:
            history["doi"][doi] = record
        if title:
            history["title"][title] = record
    return history


def lookup_history(row: pd.Series, history: dict[str, dict[str, dict[str, str]]]) -> dict[str, str] | None:
    pmid = normalize_pmid(row.get("PubMed ID") or row.get("PMID"))
    doi = normalize_doi(row.get("DOI"))
    title = normalize_text(row.get("Article Title") or row.get("Title"))
    if pmid and pmid in history.get("pmid", {}):
        return history["pmid"][pmid]
    if doi and doi in history.get("doi", {}):
        return history["doi"][doi]
    if title and title in history.get("title", {}):
        return history["title"][title]
    return None


def split_categories(value: object) -> set[str]:
    if value is None or pd.isna(value):
        return set()
    text = str(value).strip().lower()
    if not text:
        return set()

    protected = {
        "allergy and immunology": "allergy_immunology",
        "development and behavior": "development_behavior",
        "general and reviews": "general_reviews",
    }
    for phrase, token in protected.items():
        text = text.replace(phrase, token)
    text = re.sub(r"\badditional\s*:\s*", ";", text)
    text = text.replace("&", ";")
    text = re.sub(r"\band\b", ";", text)
    text = text.replace(",", ";").replace("/", ";").replace("|", ";")
    for phrase, token in protected.items():
        text = text.replace(token, phrase)

    parts = [re.sub(r"\s+", " ", part).strip(" .;:?") for part in text.split(";")]
    categories: set[str] = set()
    for part in parts:
        if not part:
            continue
        if part in CATEGORY_ALIASES:
            categories.add(CATEGORY_ALIASES[part])
            continue
        for alias, category in CATEGORY_ALIASES.items():
            if re.search(rf"\b{re.escape(alias)}\b", part):
                categories.add(category)

    return {category for category in categories if category in CATEGORIES}


def has_exclusion_label(value: object) -> bool:
    if value is None or pd.isna(value):
        return False
    text = normalize_text(value)
    return any(token in text.split() for token in ("exclude", "excluded", "exclusion"))


def is_reviewed_or_viewed(row: pd.Series, history_match: dict[str, str] | None) -> bool:
    found_in_sheets = normalize_text(row.get("Found in Sheets?")) == "yes"
    found_in_zotero = normalize_text(row.get("Found in Zotero?")) == "yes"
    return found_in_sheets or found_in_zotero


def derive_human_labels(row: pd.Series, history_match: dict[str, str] | None) -> tuple[set[str], str, str]:
    if history_match is not None:
        raw_category = str(history_match.get("category", "") or "").strip()
        if raw_category:
            return split_categories(raw_category), raw_category, "Human screening history"
        if history_match.get("screen") == "out":
            return set(), "", "Human screening history (screened out)"

    raw_zotero = str(row.get("If in Zotero - Category", "") or "").strip()
    if raw_zotero and not has_exclusion_label(raw_zotero):
        return split_categories(raw_zotero), raw_zotero, "Zotero category"
    if raw_zotero and has_exclusion_label(raw_zotero):
        return set(), "", "Zotero exclusion"
    return set(), "", "No explicit human category"


def derive_ai_labels(row: pd.Series) -> tuple[set[str], str, str]:
    ai_screen = screening_direction(row.get("OpenAI Screening In or Out") or row.get("Relevance"))
    raw_category = str(row.get("OpenAI Assigned Category", "") or "").strip()
    if ai_screen == "out":
        return set(), "", "AI screened out"
    return split_categories(raw_category), raw_category, "AI category"


def build_evaluation_dataset(review_df: pd.DataFrame, history: dict[str, dict[str, dict[str, str]]]) -> pd.DataFrame:
    rows = []
    for _, row in review_df.iterrows():
        history_match = lookup_history(row, history)
        if not is_reviewed_or_viewed(row, history_match):
            continue

        human_labels, human_raw, human_source = derive_human_labels(row, history_match)
        ai_labels, ai_raw, ai_source = derive_ai_labels(row)
        rows.append(
            {
                "PubMed ID": row.get("PubMed ID", ""),
                "Article Title": row.get("Article Title", ""),
                "Date": row.get("Date", ""),
                "Found in Sheets?": row.get("Found in Sheets?", ""),
                "Found in Zotero?": row.get("Found in Zotero?", ""),
                "Human Label Source": human_source,
                "AI Label Source": ai_source,
                "Human Raw Labels": human_raw,
                "AI Raw Labels": ai_raw,
                "Human Normalized Labels": "; ".join(category for category in CATEGORIES if category in human_labels),
                "AI Normalized Labels": "; ".join(category for category in CATEGORIES if category in ai_labels),
                "Human Label Count": len(human_labels),
                "AI Label Count": len(ai_labels),
            }
        )
    return pd.DataFrame(rows)


def categories_in_dataset(details: pd.DataFrame) -> list[str]:
    seen: set[str] = set()
    for _, row in details.iterrows():
        seen.update(split_categories(row.get("Human Raw Labels")))
        seen.update(split_categories(row.get("AI Raw Labels")))
    return [category for category in CATEGORIES if category in seen]


def cohen_kappa_from_counts(tp: int, fp: int, fn: int, tn: int) -> float | str:
    total = tp + fp + fn + tn
    if total == 0:
        return ""
    observed = (tp + tn) / total
    human_positive = (tp + fn) / total
    human_negative = (fp + tn) / total
    ai_positive = (tp + fp) / total
    ai_negative = (fn + tn) / total
    expected = (human_positive * ai_positive) + (human_negative * ai_negative)
    if expected == 1:
        return ""
    return (observed - expected) / (1 - expected)


def build_category_metrics(details: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for category in categories_in_dataset(details):
        tp = fp = fn = tn = 0
        for _, row in details.iterrows():
            human_labels = split_categories(row.get("Human Raw Labels"))
            ai_labels = split_categories(row.get("AI Raw Labels"))
            human_yes = category in human_labels
            ai_yes = category in ai_labels
            if human_yes and ai_yes:
                tp += 1
            elif ai_yes and not human_yes:
                fp += 1
            elif human_yes and not ai_yes:
                fn += 1
            else:
                tn += 1
        precision = tp / (tp + fp) if tp + fp else ""
        recall = tp / (tp + fn) if tp + fn else ""
        f1 = 2 * precision * recall / (precision + recall) if precision != "" and recall != "" and precision + recall else ""
        kappa = cohen_kappa_from_counts(tp, fp, fn, tn)
        rows.append(
            {
                "Category": category,
                "TP": tp,
                "FP": fp,
                "FN": fn,
                "TN": tn,
                "Precision": precision,
                "Recall": recall,
                "F1 Score": f1,
                "Cohen's Kappa": kappa,
            }
        )
    return pd.DataFrame(rows)


def build_overall_metrics(category_summary: pd.DataFrame) -> pd.DataFrame:
    if category_summary.empty:
        return pd.DataFrame(
            [
                {"Average": "Micro", "Precision": "", "Recall": "", "F1 Score": "", "Cohen's Kappa": ""},
                {"Average": "Macro", "Precision": "", "Recall": "", "F1 Score": "", "Cohen's Kappa": ""},
            ]
        )

    tp = int(category_summary["TP"].sum())
    fp = int(category_summary["FP"].sum())
    fn = int(category_summary["FN"].sum())
    tn = int(category_summary["TN"].sum()) if "TN" in category_summary.columns else 0

    micro_precision = tp / (tp + fp) if tp + fp else ""
    micro_recall = tp / (tp + fn) if tp + fn else ""
    micro_f1 = (
        2 * micro_precision * micro_recall / (micro_precision + micro_recall)
        if micro_precision != "" and micro_recall != "" and micro_precision + micro_recall
        else ""
    )
    micro_kappa = cohen_kappa_from_counts(tp, fp, fn, tn) if "TN" in category_summary.columns else ""

    precision_values = category_summary["Precision"].apply(lambda value: None if value == "" else float(value)).dropna()
    recall_values = category_summary["Recall"].apply(lambda value: None if value == "" else float(value)).dropna()
    f1_values = category_summary["F1 Score"].apply(lambda value: None if value == "" else float(value)).dropna()
    kappa_values = category_summary["Cohen's Kappa"].apply(lambda value: None if value == "" else float(value)).dropna()

    macro_precision = float(precision_values.mean()) if not precision_values.empty else ""
    macro_recall = float(recall_values.mean()) if not recall_values.empty else ""
    macro_f1 = float(f1_values.mean()) if not f1_values.empty else ""
    macro_kappa = float(kappa_values.mean()) if not kappa_values.empty else ""

    return pd.DataFrame(
        [
            {"Average": "Micro", "Precision": micro_precision, "Recall": micro_recall, "F1 Score": micro_f1, "Cohen's Kappa": micro_kappa},
            {"Average": "Macro", "Precision": macro_precision, "Recall": macro_recall, "F1 Score": macro_f1, "Cohen's Kappa": macro_kappa},
        ]
    )


def build_run_log(
    input_path: Path,
    output_path: Path,
    history_source: str | None,
    review_df: pd.DataFrame,
    details: pd.DataFrame,
    category_summary: pd.DataFrame,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Field": "Input workbook", "Value": str(input_path)},
            {"Field": "Input sheet", "Value": DEFAULT_SHEET},
            {"Field": "Human screening source", "Value": history_source or ""},
            {"Field": "Output workbook", "Value": str(output_path)},
            {"Field": "Generated", "Value": time.ctime()},
            {"Field": "Scoring approach", "Value": "Reviewed/viewed articles only; multilabel category scoring; Exclusion removed; General+Reviews combined."},
            {"Field": "Reviewed comparison rows", "Value": len(review_df)},
            {"Field": "Evaluation dataset rows", "Value": len(details)},
            {"Field": "Categories scored", "Value": len(category_summary)},
            {"Field": "Review filter", "Value": "Includes only rows from Review_Comparison that were actually reviewed/viewed via human screening history or Zotero presence/category."},
            {"Field": "Review filter exact rule", "Value": 'An article is scored only when "Found in Sheets?" = Yes or "Found in Zotero?" = Yes.'},
            {"Field": "Missing article rule", "Value": 'Articles without "Found in Sheets?" = Yes and without "Found in Zotero?" = Yes are excluded from evaluation and do not create FP/FN.'},
            {"Field": "Normalization", "Value": 'Combines "General" and "Reviews" into "General and Reviews"; ignores "Additional:"; standardizes delimiters; removes Exclusion from scoring.'},
            {"Field": "Multilabel scoring", "Value": "Compares categories independently per article rather than requiring exact full-string label matches."},
            {"Field": "Inference rule", "Value": "Uses only explicit human and AI labels present in the reviewed comparison workflow; does not infer missing categories."},
        ]
    )


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    match_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    mismatch_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
        headers = [cell.value for cell in ws[1]]
        max_row = max(ws.max_row, 2)

        if "Category" in headers:
            col = headers.index("Category") + 1
            for row in range(2, max_row + 1):
                category = str(ws.cell(row=row, column=col).value or "")
                color = CATEGORY_COLORS.get(category, "FFFFFF")
                ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for header in ("Human Normalized Labels", "AI Normalized Labels"):
            if header in headers:
                letter = get_column_letter(headers.index(header) + 1)
                compare_col = get_column_letter(headers.index("Human Normalized Labels") + 1) if "Human Normalized Labels" in headers else None
                ai_col = get_column_letter(headers.index("AI Normalized Labels") + 1) if "AI Normalized Labels" in headers else None
                if compare_col and ai_col:
                    cell_range = f"{letter}2:{letter}{max_row}"
                    ws.conditional_formatting.add(
                        cell_range,
                        FormulaRule(formula=[f'${compare_col}2=${ai_col}2'], fill=match_fill),
                    )
                    ws.conditional_formatting.add(
                        cell_range,
                        FormulaRule(formula=[f'AND(${compare_col}2<>"",${ai_col}2<>"",${compare_col}2<>${ai_col}2)'], fill=mismatch_fill),
                    )

        widths = {
            "PubMed ID": 12,
            "Article Title": 70,
            "Date": 14,
            "Human Raw Labels": 32,
            "AI Raw Labels": 32,
            "Human Normalized Labels": 40,
            "AI Normalized Labels": 40,
            "Human Label Source": 26,
            "AI Label Source": 18,
            "Field": 28,
            "Value": 90,
            "Average": 14,
            "Category": 30,
        }
        for index, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(index)].width = widths.get(str(header), 18)
    wb.save(path)


def write_workbook(
    output_path: Path,
    verification: pd.DataFrame,
    category_summary: pd.DataFrame,
    overall_metrics: pd.DataFrame,
    run_log: pd.DataFrame,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        verification.to_excel(writer, sheet_name="Evaluation_Dataset", index=False)
        category_summary.to_excel(writer, sheet_name="Category_Summary", index=False)
        overall_metrics.to_excel(writer, sheet_name="Overall_Metrics", index=False)
        run_log.to_excel(writer, sheet_name="Run_Log", index=False)
    format_workbook(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Calculate reviewed-only multilabel performance metrics.")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Review comparison workbook to analyze.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Metrics workbook to create.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Sheet in the review workbook to read.")
    parser.add_argument(
        "--screening-history",
        default=DEFAULT_SCREENING_HISTORY_URL,
        help="Human screening .xlsx/.csv/.tsv file or accessible Google Sheets URL.",
    )
    return parser.parse_args()


def print_verification(details: pd.DataFrame) -> None:
    print("Evaluation dataset (reviewed/viewed articles only):")
    for _, row in details.iterrows():
        print(
            f"{row.get('PubMed ID', '')}\t{row.get('Article Title', '')}\t"
            f"Human=[{row.get('Human Normalized Labels', '')}]\t"
            f"AI=[{row.get('AI Normalized Labels', '')}]"
        )


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    review_df = pd.read_excel(input_path, sheet_name=args.sheet, dtype=str)
    history = load_human_history(args.screening_history)
    verification = build_evaluation_dataset(review_df, history)
    category_summary = build_category_metrics(verification)
    overall_metrics = build_overall_metrics(category_summary)
    run_log = build_run_log(input_path, output_path, args.screening_history, review_df, verification, category_summary)

    print_verification(verification)
    write_workbook(output_path, verification, category_summary, overall_metrics, run_log)
    print(f"Metrics workbook written: {output_path.resolve()}")


if __name__ == "__main__":
    main()
