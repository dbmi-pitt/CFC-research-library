"""Sort a chosen list of Zotero articles without using current Zotero folders.

This script reads Zotero item keys from a text file, fetches each item and any
indexed full text from Zotero, asks OpenAI to classify using the project
criteria, and writes a review workbook. It intentionally ignores collection
membership so existing folder placement cannot influence the recommendation.

Run from the project folder:
    uv run python sort_zotero_articles_without_folders.py
"""

from __future__ import annotations

import argparse
import json
import os
import textwrap
import time
from pathlib import Path
from urllib.parse import quote

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cfc_research_library import CATEGORY_COLORS, SECTIONS
from generate_unfiled_zotero_article_assignments import (
    DEFAULT_OPENAI_MODEL,
    EXCLUDED_CATEGORIES,
    allowed_categories,
    author_text,
    clean_note,
    display_title,
    extract_pmid,
    full_text_exclusion_reason,
    item_text,
    load_criteria_text,
    load_env_file,
    parse_json_response,
    require_env,
    year_from_date,
    zotero_api_get_json,
    zotero_fulltext_for_item,
)


DEFAULT_INPUT = "zotero_article_sort_input.txt"
DEFAULT_OUTPUT = "reports/Zotero_Article_Sort_No_Folder_Lookup.xlsx"


def parse_input_list(path: Path) -> list[dict[str, str]]:
    rows = []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = line.strip()
        if not line:
            continue
        parts = line.split(maxsplit=1)
        key = parts[0].strip()
        title = parts[1].strip().strip("[]") if len(parts) > 1 else ""
        rows.append({"Zotero Item Key": key, "Input Title": title})
    return rows


def fetch_items_without_collections(input_rows: list[dict[str, str]]) -> pd.DataFrame:
    group_id = require_env("ZOTERO_GROUP_ID")
    api_key = require_env("ZOTERO_API_KEY")
    rows = []

    for input_row in input_rows:
        key = input_row["Zotero Item Key"]
        item = zotero_api_get_json(f"https://api.zotero.org/groups/{group_id}/items/{quote(key)}", api_key)
        data = item.get("data", {}) if isinstance(item, dict) else {}

        title = display_title(data, key) if data else input_row["Input Title"]
        full_text, full_text_source = zotero_fulltext_for_item(group_id, api_key, key) if data else ("", "")
        abstract = data.get("abstractNote", "") or clean_note(data.get("note", ""))
        text_for_ai = "\n\n".join(
            part
            for part in [
                item_text(data, title) if data else title,
                f"ZOTERO INDEXED FULL TEXT:\n{full_text[:18000]}" if full_text else "",
            ]
            if part
        )

        rows.append(
            {
                "Zotero Item Key": key,
                "Input Title": input_row["Input Title"],
                "Title": title,
                "Year": year_from_date(data.get("date", "")) if data else "",
                "Authors": author_text(data.get("creators")) if data else "",
                "Journal / Publication": data.get("publicationTitle", "")
                or data.get("proceedingsTitle", "")
                or data.get("bookTitle", ""),
                "DOI": data.get("DOI", ""),
                "PMID": extract_pmid(data.get("extra", "")),
                "Zotero Item Found?": "Yes" if data else "No",
                "Full Text Found?": "Yes" if full_text else "No",
                "Full Text Source": full_text_source,
                "Abstract": abstract,
                "_Text_For_AI": text_for_ai,
            }
        )

    return pd.DataFrame(rows).fillna("")


def classify_batch(client, batch: pd.DataFrame, model_name: str, criteria_text: str) -> list[dict]:
    articles = []
    for idx, row in batch.iterrows():
        articles.append(
            {
                "row_index": int(idx),
                "zotero_item_key": str(row.get("Zotero Item Key", "")),
                "title": str(row.get("Title", "")),
                "input_title": str(row.get("Input Title", "")),
                "abstract": str(row.get("Abstract", ""))[:5000],
                "journal": str(row.get("Journal / Publication", "")),
                "year": str(row.get("Year", "")),
                "doi": str(row.get("DOI", "")),
                "pmid": str(row.get("PMID", "")),
                "full_text_found": str(row.get("Full Text Found?", "")),
                "article_text": str(row.get("_Text_For_AI", ""))[:20000],
            }
        )

    prompt = textwrap.dedent(
        f"""
        Classify each Zotero article for a CFC syndrome research library.

        Important: do not use or infer current Zotero folder placement. Current folders are intentionally hidden.
        Sort based only on the title, metadata, abstract, and Zotero indexed full text supplied here.

        Apply the inclusion/exclusion criteria carefully.

        Output rules:
        - Exclude articles that are clearly not related to CFC, CFC-associated RAS/MAPK biology, or clinically relevant RASopathy comparison.
        - Exclude Noonan-only, Costello-only, Legius-only, NF1-only, cancer-only, somatic mutation-only, or broad pathway papers when they do not provide CFC data, a CFC-associated mutation/model, a shared RASopathy mechanism, or useful clinical comparison for CFC.
        - Inspect the supplied full text when available. If CFC is only barely mentioned, appears only in a passing disease list, appears only in a figure/table label, appears only in references, or is not substantively discussed in the aims/results/discussion, set decision to "Excluded".
        - If the article is clearly about another condition and only mentions RAS/MAPK, BRAF, MAP2K1, MAP2K2, KRAS, or RASopathy in passing, set decision to "Excluded".
        - If CFC relevance is uncertain but plausible, set decision to "Needs review" and assign the best tentative category so Lexi can check it.
        - If an article is about Noonan syndrome, Costello syndrome, Legius syndrome, NF1, or general RASopathies but may provide useful comparison, shared mechanism, Noonan-spectrum context, or CFC-relevant mutation/model information, use "Needs review" rather than excluding it.
        - Do not exclude CFC-specific clinical guidelines, diagnostic criteria papers, CFC index papers, prenatal diagnosis overviews, or CFC-focused management papers just because they are not original research. Classify these as "General and Reviews".
        - If a RASopathy-wide review or overview meaningfully discusses CFC, compares CFC with other RASopathies, or provides useful CFC clinical/genetic/management context, classify it as "General and Reviews" instead of excluding it.
        - If the article is a review, broad overview, clinical guideline, consensus statement, diagnostic summary, educational summary, or case report/case series that does not present a new mutation, new disease mechanism, new experimental result, new genotype-phenotype analysis, or previously undescribed clinical manifestation, classify it as "General and Reviews".
        - If a review mainly addresses treatment options or therapeutic landscape, prefer "Treatments" as the primary category and add the organ system as an additional category when appropriate.
        - Human review examples to follow: CFC clinical management guidelines, CFC index/diagnostic criteria papers, prenatal diagnosis overviews with meaningful CFC discussion, and RASopathy reviews with substantive CFC comparison belong in "General and Reviews"; autoimmune or other RASopathy cohorts with actual CFC participants can belong in the relevant specialty; papers that only allude to CFC in one figure/table/reference should be excluded.
        - If the article presents original CFC-specific evidence, choose the most specific clinical/scientific category.
        - If a paper primarily identifies or validates a CFC pathogenic variant, genotype-phenotype relationship, inheritance pattern, or variant function for clinical genetics, prefer "Genetics".
        - If a paper primarily studies hypertrophic cardiomyopathy, heart function, MEK inhibition for cardiac disease, arrhythmia, or cardiac surveillance, prefer "Cardiology".
        - If a paper is primarily epilepsy, EEG, seizure phenotype, or seizure treatment, prefer "Seizures" rather than Neurology.
        - If a paper is primarily broader motor milestones, behavior, cognition, motor function, hypotonia, or neuroimaging without seizure focus, prefer "Neurology" or "Development and Behavior" based on the main outcome.
        - If a paper mainly describes skin, hair, nail, eczema, keratosis, or other cutaneous findings, prefer "Dermatology".
        - If excluded, leave ai_category blank and explain the specific exclusion reason.
        - If decision is "Needs review", still provide the best tentative ai_category.
        - If included, choose one primary category in ai_category.
        - If the article clearly belongs in more than one category, mention additional categories in additional_categories.
        - Do not add multiple categories just because the article mentions multiple symptoms.
        - Do not use Historical Articles or Conferences.

        Inclusion/exclusion criteria and category definitions:
        {criteria_text}

        Articles:
        {json.dumps(articles, indent=2)}

        Return only valid JSON in this exact shape:
        {{
          "articles": [
            {{
              "row_index": 0,
              "decision": "Included",
              "ai_category": "Cardiology",
              "additional_categories": ["Genetics"],
              "confidence": "High",
              "reasoning": "Brief rationale based on the full text/abstract."
            }}
          ]
        }}
        """
    ).strip()

    response = client.responses.create(model=model_name, input=prompt)
    return parse_json_response(response.output_text).get("articles", [])


def add_ai_sorting(df: pd.DataFrame, model_name: str, batch_size: int, criteria_text: str) -> pd.DataFrame:
    from openai import OpenAI

    output = df.copy()
    output["AI Decision"] = ""
    output["AI Categorization"] = ""
    output["AI Confidence"] = ""
    output["AI Reasoning"] = ""
    client = OpenAI(api_key=require_env("OPENAI_API_KEY"))

    for start in range(0, len(output), batch_size):
        batch = output.iloc[start : start + batch_size]
        try:
            assignments = classify_batch(client, batch, model_name, criteria_text)
        except Exception as exc:
            assignments = []
            for idx, row in batch.iterrows():
                reason = full_text_exclusion_reason(row.get("_Text_For_AI", ""))
                output.at[idx, "AI Decision"] = "Excluded" if reason else "Needs review"
                output.at[idx, "AI Confidence"] = "Low"
                output.at[idx, "AI Reasoning"] = f"{reason} OpenAI batch failed. Error: {type(exc).__name__}".strip()

        for item in assignments:
            idx = item.get("row_index")
            if idx not in output.index:
                continue
            decision = str(item.get("decision", "") or "Needs review").strip()
            category = str(item.get("ai_category", "") or "").strip()
            reason = full_text_exclusion_reason(output.at[idx, "_Text_For_AI"])
            if reason:
                decision = "Excluded"
                category = ""
            if category in EXCLUDED_CATEGORIES or category not in SECTIONS:
                category = ""

            additional = item.get("additional_categories", [])
            if isinstance(additional, list):
                additional = [
                    str(value).strip()
                    for value in additional
                    if str(value).strip() in SECTIONS and str(value).strip() not in EXCLUDED_CATEGORIES
                ]
            else:
                additional = []

            category_text = category
            if category and additional:
                category_text = f"{category}; Additional: {', '.join(additional)}"

            reasoning = str(item.get("reasoning", "") or item.get("rationale", "") or "").strip()
            if reason:
                reasoning = f"{reason} OpenAI rationale: {reasoning}".strip()

            output.at[idx, "AI Decision"] = decision
            output.at[idx, "AI Categorization"] = category_text
            output.at[idx, "AI Confidence"] = str(item.get("confidence", "") or "").strip().replace("Moderate", "Medium")
            output.at[idx, "AI Reasoning"] = reasoning

        for idx, row in batch.iterrows():
            if str(output.at[idx, "AI Decision"]).strip():
                continue
            reason = full_text_exclusion_reason(row.get("_Text_For_AI", ""))
            output.at[idx, "AI Decision"] = "Excluded" if reason else "Needs review"
            output.at[idx, "AI Confidence"] = "Low"
            output.at[idx, "AI Reasoning"] = (
                reason
                or "OpenAI did not return a row-level assignment for this article; marked for human review."
            )

    return output


def build_report(df: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Zotero Item Key": df["Zotero Item Key"],
            "Title": df["Title"],
            "Year": df["Year"],
            "Zotero Item Found?": df["Zotero Item Found?"],
            "Full Text Found?": df["Full Text Found?"],
            "AI Decision": df["AI Decision"],
            "AI Categorization": df["AI Categorization"],
            "AI Confidence": df["AI Confidence"],
            "AI Reasoning": df["AI Reasoning"],
            "Lexi Category Assignment": "",
            "Lexi Notes": "",
            "PMID": df["PMID"],
            "DOI": df["DOI"],
        }
    )


def criteria_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Category": section.name,
                "Description": section.description,
                "Inclusion": section.inclusion,
                "Exclusion": section.exclusion,
            }
            for name, section in SECTIONS.items()
            if name not in EXCLUDED_CATEGORIES
        ]
    )


def write_workbook(path: Path, report: pd.DataFrame, model_name: str, criteria_text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    excluded = report[report["AI Decision"].str.lower().eq("excluded")].copy()
    summary = report.groupby(["AI Decision", "AI Categorization"], dropna=False).size().reset_index(name="Article Count")
    run_log = pd.DataFrame(
        [
            {"Metric": "Articles sorted", "Value": len(report)},
            {"Metric": "Excluded rows", "Value": len(excluded)},
            {"Metric": "OpenAI model", "Value": model_name},
            {"Metric": "Current Zotero folders used?", "Value": "No"},
            {"Metric": "Generated", "Value": time.ctime()},
        ]
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="AI_Sorted", index=False)
        excluded.to_excel(writer, sheet_name="Excluded", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        criteria_frame().to_excel(writer, sheet_name="Criteria", index=False)
        pd.DataFrame({"Criteria Used": [criteria_text]}).to_excel(writer, sheet_name="Full_Criteria_Text", index=False)
        run_log.to_excel(writer, sheet_name="Run_Log", index=False)

    format_workbook(path)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    widths = {
        "Zotero Item Key": 16,
        "Title": 72,
        "Year": 10,
        "Zotero Item Found?": 18,
        "Full Text Found?": 18,
        "AI Decision": 16,
        "AI Categorization": 38,
        "AI Confidence": 16,
        "AI Reasoning": 72,
        "Lexi Category Assignment": 28,
        "Lexi Notes": 42,
        "PMID": 16,
        "DOI": 28,
    }

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        headers = [cell.value for cell in sheet[1]]
        max_row = max(sheet.max_row, 2)
        if sheet_name in {"AI_Sorted", "Excluded"} and "Lexi Category Assignment" in headers:
            col = get_column_letter(headers.index("Lexi Category Assignment") + 1)
            validation = DataValidation(
                type="list",
                formula1=f"='Criteria'!$A$2:$A${len(allowed_categories()) + 1}",
                allow_blank=True,
            )
            sheet.add_data_validation(validation)
            validation.add(f"{col}2:{col}{max_row}")

        if sheet_name in {"AI_Sorted", "Excluded"}:
            for category_header in ("AI Categorization", "Lexi Category Assignment"):
                if category_header in headers:
                    col_index = headers.index(category_header) + 1
                    for row_num in range(2, max_row + 1):
                        value = str(sheet.cell(row=row_num, column=col_index).value or "")
                        category = value.split(";")[0].strip()
                        color = CATEGORY_COLORS.get(category, "FFFFFF")
                        sheet.cell(row=row_num, column=col_index).fill = PatternFill(
                            start_color=color,
                            end_color=color,
                            fill_type="solid",
                        )

        for idx, header in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = widths.get(
                str(header),
                42 if sheet_name in {"Criteria", "Full_Criteria_Text"} else 24,
            )
            if str(header) in {"Title", "AI Reasoning", "Lexi Notes", "Criteria Used"}:
                for row_num in range(2, min(max_row, 250) + 1):
                    sheet.cell(row=row_num, column=idx).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sort selected Zotero articles without using current folder placement.")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Text file with Zotero item keys and optional titles.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Workbook output path.")
    parser.add_argument("--openai-model", default=os.getenv("OPENAI_CATEGORY_MODEL", DEFAULT_OPENAI_MODEL))
    parser.add_argument("--batch-size", type=int, default=8)
    parser.add_argument("--criteria-file", type=Path, default=None)
    return parser.parse_args()


def main() -> None:
    load_env_file()
    args = parse_args()
    input_rows = parse_input_list(Path(args.input))
    criteria_text = load_criteria_text(args.criteria_file)
    raw = fetch_items_without_collections(input_rows)
    sorted_articles = add_ai_sorting(raw, args.openai_model, args.batch_size, criteria_text)
    report = build_report(sorted_articles)
    write_workbook(Path(args.output), report, args.openai_model, criteria_text)
    print(f"Workbook written: {Path(args.output).resolve()}")


if __name__ == "__main__":
    main()
