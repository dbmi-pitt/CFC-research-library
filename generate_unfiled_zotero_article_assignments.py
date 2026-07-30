"""Generate a clean Excel workbook for unfiled Zotero full-text assignments.

This script pulls unfiled, article-like records from Zotero, tries to use
Zotero indexed full text plus the title/abstract, asks OpenAI to apply the
new CFC inclusion/exclusion criteria, and creates a workbook with:

    - All_Unfiled
    - Excluded
    - Criteria
    - Run_Log

The output name is intentionally different from prior workbooks so it will not
overwrite edited files.

Run from the project folder:
    uv run python generate_unfiled_zotero_article_assignments.py
"""

from __future__ import annotations

import argparse
import json
import os
import re
import textwrap
import time
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from cfc_research_library import CATEGORY_COLORS, SECTIONS


DEFAULT_OUTPUT = "reports/Unfiled_Zotero_FullText_AI_Assignment.xlsx"
DEFAULT_OPENAI_MODEL = "gpt-5.4-mini"
ARTICLE_ITEM_TYPES = {"journalArticle", "newspaperArticle", "bookSection"}
EXCLUDED_CATEGORIES = {"Historical Articles", "Conferences"}
BAD_TITLES = {"pdf", "full text pdf", "full text", "pubmed entry"}
DEFAULT_CRITERIA_FILE = Path(
    r"C:\Users\lexim\.codex\attachments\772c3eed-9d2d-47af-8885-b8f6a5fa19aa\pasted-text.txt"
)

CFC_SPECIFIC_TERMS = (
    "cardiofaciocutaneous",
    "cardio-facio-cutaneous",
    "cardio facio cutaneous",
    "cfc syndrome",
    "cfcs",
)

NON_CFC_RASOPATHY_TERMS = (
    "noonan syndrome",
    "noonan-spectrum",
    "noonan spectrum",
    "costello syndrome",
    "legius syndrome",
    "neurofibromatosis type 1",
)

ZOTERO_CATEGORY_ALIASES = {
    "general_and_reviews": "General and Reviews",
    "general and reviews": "General and Reviews",
    "research studies": "Research Studies",
    "allergy and immunology untitled": "Allergy and Immunology",
    "exclusion": "Excluded",
}

MAX_ZOTERO_EXAMPLES_PER_CATEGORY = 5


def load_env_file(path: Path = Path(".env")) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def extract_pmid(extra: object) -> str:
    if not extra:
        return ""
    match = re.search(r"\bPMID:\s*(\d+)\b", str(extra), flags=re.IGNORECASE)
    return match.group(1) if match else ""


def author_text(creators: list[dict] | None) -> str:
    names = []
    for creator in creators or []:
        if creator.get("creatorType") not in {"author", "editor"}:
            continue
        name = " ".join(
            part for part in [creator.get("firstName", ""), creator.get("lastName", "")] if part
        ).strip()
        if not name:
            name = creator.get("name", "")
        if name:
            names.append(name)
    return ", ".join(names)


def year_from_date(value: object) -> str:
    if not value:
        return ""
    match = re.search(r"\b(19\d{2}|20\d{2})\b", str(value))
    return match.group(1) if match else ""


def clean_note(value: object) -> str:
    text = re.sub(r"<[^>]+>", " ", str(value or ""))
    return re.sub(r"\s+", " ", text).strip()


def normalize_zotero_category(value: object) -> str:
    text = re.sub(r"\s+", " ", str(value or "").replace("_", " ")).strip()
    key = text.lower()
    if key in ZOTERO_CATEGORY_ALIASES:
        return ZOTERO_CATEGORY_ALIASES[key]
    for category in SECTIONS:
        if category.lower() == key:
            return category
    return text


def display_title(data: dict, key: str) -> str:
    for field in ("title", "filename", "publicationTitle", "bookTitle", "proceedingsTitle"):
        value = str(data.get(field, "") or "").strip()
        if value:
            return value
    note = clean_note(data.get("note", ""))
    if note:
        return note[:120]
    return f"Untitled {data.get('itemType', 'item')} ({key})"


def item_text(data: dict, title: str) -> str:
    tags = " ".join(tag.get("tag", "") for tag in data.get("tags", []) if isinstance(tag, dict))
    parts = [
        title,
        data.get("abstractNote", ""),
        clean_note(data.get("note", "")),
        data.get("publicationTitle", ""),
        data.get("extra", ""),
        tags,
    ]
    return ". ".join(str(part) for part in parts if part)


def short_abstract(data: dict, limit: int = 450) -> str:
    text = clean_note(data.get("abstractNote", ""))
    if not text:
        text = clean_note(data.get("note", ""))
    return text[:limit]


def full_text_exclusion_reason(text: object) -> str:
    """Return an exclusion reason only when the available text is clearly out of scope."""
    lowered = str(text or "").lower()
    has_cfc = any(term in lowered for term in CFC_SPECIFIC_TERMS)
    has_rasopathy = "rasopath" in lowered
    has_cfc_gene = any(gene in lowered for gene in ("braf", "map2k1", "map2k2", "kras"))
    has_related_rasopathy = any(term in lowered for term in NON_CFC_RASOPATHY_TERMS)

    if not has_cfc and not has_rasopathy and not has_cfc_gene and not has_related_rasopathy:
        return "Excluded by title/full-text screen: no CFC, RASopathy, or CFC-associated RAS/MAPK signal was found."

    return ""


def parse_json_response(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.IGNORECASE).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start : end + 1])
        raise


def allowed_categories() -> list[str]:
    return [category for category in SECTIONS if category not in EXCLUDED_CATEGORIES]


def load_criteria_text(criteria_file: Path | None = None) -> str:
    if criteria_file and criteria_file.exists():
        return criteria_file.read_text(encoding="utf-8", errors="replace")
    if DEFAULT_CRITERIA_FILE.exists():
        return DEFAULT_CRITERIA_FILE.read_text(encoding="utf-8", errors="replace")
    criteria_rows = []
    for name, section in SECTIONS.items():
        if name in EXCLUDED_CATEGORIES:
            continue
        criteria_rows.append(
            f"{section.name}\n"
            f"Description: {section.description}\n"
            f"Inclusion: {section.inclusion}\n"
            f"Exclusion: {section.exclusion}"
        )
    return "\n\n".join(criteria_rows)


def zotero_api_get_json(url: str, api_key: str) -> dict | list:
    request = Request(
        url,
        headers={"Zotero-API-Key": api_key, "Accept": "application/json"},
    )
    try:
        with urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception:
        return {}


def zotero_fulltext_for_item(group_id: str, api_key: str, item_key: str) -> tuple[str, str]:
    encoded_key = quote(item_key)
    base = f"https://api.zotero.org/groups/{group_id}/items/{encoded_key}"
    texts = []
    sources = []

    for key, label in [(item_key, "item")]:
        url = f"https://api.zotero.org/groups/{group_id}/items/{quote(key)}/fulltext"
        data = zotero_api_get_json(url, api_key)
        if isinstance(data, dict) and data.get("content"):
            texts.append(str(data.get("content", "")))
            sources.append(label)

    children = zotero_api_get_json(f"{base}/children?limit=100", api_key)
    if isinstance(children, list):
        for child in children:
            child_data = child.get("data", {})
            if child_data.get("itemType") != "attachment":
                continue
            child_key = child.get("key") or child_data.get("key")
            if not child_key:
                continue
            fulltext = zotero_api_get_json(
                f"https://api.zotero.org/groups/{group_id}/items/{quote(child_key)}/fulltext",
                api_key,
            )
            if isinstance(fulltext, dict) and fulltext.get("content"):
                texts.append(str(fulltext.get("content", "")))
                sources.append(f"attachment:{child_key}")

    combined = "\n\n".join(text for text in texts if text).strip()
    return combined, ", ".join(sources)


def fetch_zotero_assignment_examples() -> str:
    """Collect examples from already-filed Zotero items to mirror Lexi's final folders."""
    from pyzotero import zotero

    group_id = require_env("ZOTERO_GROUP_ID")
    api_key = require_env("ZOTERO_API_KEY")
    zot = zotero.Zotero(group_id, "group", api_key)
    collections = {
        collection["key"]: collection["data"].get("name", "")
        for collection in zot.everything(zot.collections())
    }
    examples: dict[str, list[str]] = {}
    excluded_examples: list[str] = []

    for collection_key, collection_name in collections.items():
        category = normalize_zotero_category(collection_name)
        if category != "Excluded" and (category in EXCLUDED_CATEGORIES or category not in SECTIONS):
            continue

        encoded_collection_key = quote(collection_key)
        url = (
            f"https://api.zotero.org/groups/{group_id}/collections/"
            f"{encoded_collection_key}/items?limit=25"
        )
        collection_items = zotero_api_get_json(url, api_key)
        if not isinstance(collection_items, list):
            continue

        for item in collection_items:
            data = item.get("data", {})
            item_type = data.get("itemType", "")
            if item_type in {"attachment", "note", "annotation"}:
                continue
            title = display_title(data, item.get("key", ""))
            if not title or title.strip().lower() in BAD_TITLES:
                continue

            example_text = title
            abstract = short_abstract(data)
            if abstract:
                example_text = f"{title} -- {abstract}"

            if category == "Excluded":
                if len(excluded_examples) < MAX_ZOTERO_EXAMPLES_PER_CATEGORY:
                    excluded_examples.append(example_text)
                break

            examples.setdefault(category, [])
            if len(examples[category]) < MAX_ZOTERO_EXAMPLES_PER_CATEGORY:
                examples[category].append(example_text)
            if len(examples[category]) >= MAX_ZOTERO_EXAMPLES_PER_CATEGORY:
                break

    lines = [
        "Use these existing Zotero folder assignments as examples of Lexi's final categorization style.",
        "These examples should guide tie-breaks when the written criteria overlap.",
        "Do not copy a category blindly; apply the same reasoning pattern to the new article.",
    ]
    if excluded_examples:
        lines.append("\nExcluded examples from Zotero:")
        for text in excluded_examples:
            lines.append(f"- Excluded: {text[:700]}")
    for category in allowed_categories():
        category_examples = examples.get(category, [])
        if not category_examples:
            continue
        lines.append(f"\n{category} examples from Zotero:")
        for text in category_examples:
            lines.append(f"- {text[:700]}")
    return "\n".join(lines)


def fetch_unfiled_articles() -> pd.DataFrame:
    from pyzotero import zotero

    group_id = require_env("ZOTERO_GROUP_ID")
    api_key = require_env("ZOTERO_API_KEY")
    zot = zotero.Zotero(group_id, "group", api_key)
    rows = []
    for item in zot.everything(zot.items()):
        data = item.get("data", {})
        item_type = data.get("itemType", "")
        if data.get("collections"):
            continue
        if item_type not in ARTICLE_ITEM_TYPES:
            continue
        key = item.get("key", "")
        title = display_title(data, key)
        if title.strip().lower() in BAD_TITLES:
            continue
        full_text, full_text_source = zotero_fulltext_for_item(group_id, api_key, key)
        rows.append(
            {
                "Zotero Item Key": key,
                "Title": title,
                "Authors": author_text(data.get("creators")),
                "Year": year_from_date(data.get("date", "")),
                "Journal / Publication": data.get("publicationTitle", "")
                or data.get("proceedingsTitle", "")
                or data.get("bookTitle", ""),
                "DOI": data.get("DOI", ""),
                "PMID": extract_pmid(data.get("extra")),
                "Item Type": item_type,
                "Abstract": data.get("abstractNote", "") or clean_note(data.get("note", "")),
                "Full Text Found?": "Yes" if full_text else "No",
                "Full Text Source": full_text_source,
                "URL": data.get("url", ""),
                "_Text_For_AI": "\n\n".join(
                    part
                    for part in [
                        item_text(data, title),
                        f"ZOTERO INDEXED FULL TEXT:\n{full_text[:18000]}" if full_text else "",
                    ]
                    if part
                ),
            }
        )
    return pd.DataFrame(rows).fillna("")


def classify_openai_batch(
    client,
    batch: pd.DataFrame,
    model_name: str,
    criteria_text: str,
    zotero_examples_text: str,
) -> list[dict]:
    articles = []
    for idx, row in batch.iterrows():
        articles.append(
            {
                "row_index": int(idx),
                "zotero_item_key": str(row.get("Zotero Item Key", "")),
                "title": str(row.get("Title", "")),
                "abstract": str(row.get("Abstract", ""))[:5000],
                "journal": str(row.get("Journal / Publication", "")),
                "year": str(row.get("Year", "")),
                "doi": str(row.get("DOI", "")),
                "pmid": str(row.get("PMID", "")),
                "full_text_found": str(row.get("Full Text Found?", "")),
                "article_text": str(row.get("_Text_For_AI", ""))[:20000],
            }
        )

    prompt = textwrap.dedent(
        f"""
        Classify each unfiled Zotero article for a CFC syndrome research library.

        Use the available Zotero indexed full text when present. If full text is not present,
        use title, abstract, journal, and metadata.

        Apply the overall inclusion/exclusion criteria and category-specific criteria carefully.

        Output rules:
        - Exclude articles that are clearly not related to CFC, CFC-associated RAS/MAPK biology, or clinically relevant RASopathy comparison.
        - Exclude Noonan-only, Costello-only, Legius-only, NF1-only, cancer-only, somatic mutation-only, or broad pathway papers when they do not provide CFC data, a CFC-associated mutation/model, a shared RASopathy mechanism, or useful clinical comparison for CFC.
        - Inspect the supplied full text when available. If CFC is only barely mentioned, appears only in a passing disease list, appears only in a figure/table label, appears only in references, or is not substantively discussed in the aims/results/discussion, set decision to "Excluded".
        - If the article is clearly about another condition and only mentions RAS/MAPK, BRAF, MAP2K1, MAP2K2, KRAS, or RASopathy in passing, set decision to "Excluded".
        - If CFC relevance is uncertain but plausible, set decision to "Needs review" and assign the best tentative category so Lexi can check it.
        - If an article is about Noonan syndrome, Costello syndrome, Legius syndrome, NF1, or general RASopathies but may provide useful comparison, shared mechanism, Noonan-spectrum context, or CFC-relevant mutation/model information, use "Needs review" rather than excluding it.
        - Do not exclude CFC-specific clinical guidelines, diagnostic criteria papers, CFC index papers, prenatal diagnosis overviews, or CFC-focused management papers just because they are not original research. Classify these as "General and Reviews".
        - If a RASopathy-wide review or overview meaningfully discusses CFC, compares CFC with other RASopathies, or provides useful CFC clinical/genetic/management context, classify it as "General and Reviews" instead of excluding it.
        - If the article is a review, broad overview, clinical guideline, consensus statement, diagnostic summary, educational summary, or case report/case series that does not present a new mutation, new disease mechanism, new experimental result, new genotype-phenotype analysis, or previously undescribed clinical manifestation, classify it as "General and Reviews".
        - If a review mainly addresses treatment options or therapeutic landscape, prefer "Treatments" as the primary category and add the organ system as an additional category when appropriate.
        - Human review examples to follow: CFC clinical management guidelines, CFC index/diagnostic criteria papers, prenatal diagnosis overviews with meaningful CFC discussion, and RASopathy reviews with substantive CFC comparison belong in "General and Reviews"; autoimmune or other RASopathy cohorts with actual CFC participants can belong in the relevant specialty; papers that only allude to CFC in one figure/table/reference should be excluded.
        - Use a specialty category only when the article contributes new original evidence or a clearly specialty-specific clinical finding for that folder.
        - Do not assign specialty folders to review articles merely because the review mentions that organ system.
        - If a paper is broad CFC clinical characterization, diagnosis, management guidance, or patient education, prefer "General and Reviews" unless there is a clearly novel specialty finding.
        - If a paper primarily identifies or validates a CFC pathogenic variant, genotype-phenotype relationship, inheritance pattern, or variant function for clinical genetics, prefer "Genetics" even if it also describes clinical features.
        - If a paper primarily studies hypertrophic cardiomyopathy, heart function, MEK inhibition for cardiac disease, arrhythmia, or cardiac surveillance, prefer "Cardiology".
        - If a paper is primarily epilepsy, EEG, seizure phenotype, or seizure treatment, prefer "Seizures" rather than Neurology.
        - If a paper is primarily broader motor milestones, behavior, cognition, motor function, hypotonia, or neuroimaging without seizure focus, prefer "Neurology" or "Development and Behavior" based on the main outcome.
        - If a paper mainly describes skin, hair, nail, eczema, keratosis, or other cutaneous findings, prefer "Dermatology" even if CFC genetics are mentioned.
        - If excluded, leave ai_category blank and explain the specific exclusion reason.
        - If decision is "Needs review", still provide the best tentative ai_category.
        - If included, choose one primary category in ai_category.
        - If the article clearly belongs in more than one category, mention additional categories in additional_categories.
        - Do not add multiple categories just because the article mentions multiple symptoms; only add them when the article substantially belongs in those sections.
        - Do not use Historical Articles or Conferences.

        New inclusion/exclusion criteria and category definitions:
        {criteria_text}

        Existing Zotero categorization examples:
        {zotero_examples_text}

        Articles:
        {json.dumps(articles, indent=2)}

        Return only valid JSON in this exact shape:
        {{
          "articles": [
            {{
              "row_index": 0,
              "decision": "Included",
              "ai_category": "Cardiology",
              "additional_categories": ["Genetics"],
              "confidence": "High",
              "reasoning": "Brief rationale based on the full text/abstract."
            }}
          ]
        }}
        """
    ).strip()

    response = client.responses.create(model=model_name, input=prompt)
    return parse_json_response(response.output_text).get("articles", [])


def add_openai_categories(
    df: pd.DataFrame,
    model_name: str,
    batch_size: int,
    criteria_text: str,
    zotero_examples_text: str,
) -> pd.DataFrame:
    from openai import OpenAI

    if df.empty:
        df["AI Decision"] = ""
        df["AI Categorization"] = ""
        df["AI Confidence"] = ""
        df["AI Reasoning"] = ""
        return df

    client = OpenAI(api_key=require_env("OPENAI_API_KEY"))
    output = df.copy()
    output["AI Decision"] = ""
    output["AI Categorization"] = ""
    output["AI Confidence"] = ""
    output["AI Reasoning"] = ""

    for start in range(0, len(output), batch_size):
        batch = output.iloc[start : start + batch_size]
        try:
            assignments = classify_openai_batch(client, batch, model_name, criteria_text, zotero_examples_text)
        except Exception as exc:
            assignments = []
            for idx, row in batch.iterrows():
                override_reason = full_text_exclusion_reason(row.get("_Text_For_AI", ""))
                output.at[idx, "AI Decision"] = "Excluded" if override_reason else "Needs review"
                output.at[idx, "AI Categorization"] = ""
                output.at[idx, "AI Confidence"] = "Low"
                if override_reason:
                    output.at[idx, "AI Reasoning"] = f"{override_reason} OpenAI batch failed. Error: {type(exc).__name__}"
                else:
                    output.at[idx, "AI Reasoning"] = f"OpenAI batch failed. Error: {type(exc).__name__}"

        for item in assignments:
            idx = item.get("row_index")
            if idx not in output.index:
                continue
            decision = str(item.get("decision", "") or "").strip() or "Needs review"
            category = str(item.get("ai_category", "") or "").strip()
            override_reason = full_text_exclusion_reason(output.at[idx, "_Text_For_AI"])
            if override_reason:
                decision = "Excluded"
                category = ""
            if category in EXCLUDED_CATEGORIES or category not in SECTIONS:
                category = ""
            additional = item.get("additional_categories", [])
            if isinstance(additional, list):
                additional = [str(value).strip() for value in additional if str(value).strip() in SECTIONS and str(value).strip() not in EXCLUDED_CATEGORIES]
            else:
                additional = []
            category_text = category
            if category and additional:
                category_text = f"{category}; Additional: {', '.join(additional)}"
            output.at[idx, "AI Decision"] = decision
            output.at[idx, "AI Categorization"] = category_text
            output.at[idx, "AI Confidence"] = str(item.get("confidence", "") or "").strip().replace("Moderate", "Medium")
            reasoning = str(item.get("reasoning", "") or item.get("rationale", "") or "").strip()
            if override_reason:
                reasoning = f"{override_reason} OpenAI rationale: {reasoning}".strip()
            output.at[idx, "AI Reasoning"] = reasoning
    return output


def build_output_frame(df: pd.DataFrame) -> pd.DataFrame:
    output = pd.DataFrame(
        {
            "Zotero Item Key": df["Zotero Item Key"],
            "Title": df["Title"],
            "AI Decision": df["AI Decision"],
            "AI Categorization": df["AI Categorization"],
            "AI Confidence": df["AI Confidence"],
            "AI Reasoning": df["AI Reasoning"],
            "Full Text Found?": df["Full Text Found?"],
            "Lexi Category Assignment": "",
            "Lexi Notes": "",
        }
    )
    return output


def criteria_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Category": section.name,
                "Description": section.description,
                "Inclusion": section.inclusion,
                "Exclusion": section.exclusion,
            }
            for name, section in SECTIONS.items()
            if name not in EXCLUDED_CATEGORIES
        ]
    )


def write_workbook(
    path: Path,
    report: pd.DataFrame,
    openai_model: str,
    criteria_text: str,
    zotero_examples_text: str,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    excluded = report[report["AI Decision"].str.lower().eq("excluded")].copy()
    summary = report.groupby(["AI Decision", "AI Categorization"], dropna=False).size().reset_index(name="Article Count")
    run_log = pd.DataFrame(
        [
            {"Metric": "Clean unfiled article rows", "Value": len(report)},
            {"Metric": "Excluded rows", "Value": len(excluded)},
            {"Metric": "OpenAI model", "Value": openai_model},
            {"Metric": "Dropdown categories excluded", "Value": ", ".join(sorted(EXCLUDED_CATEGORIES))},
            {"Metric": "Zotero examples used", "Value": "Yes"},
            {"Metric": "Generated", "Value": time.ctime()},
        ]
    )
    criteria_text_frame = pd.DataFrame({"Criteria Used": [criteria_text]})
    examples_frame = pd.DataFrame({"Zotero Assignment Examples Used": [zotero_examples_text]})

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="All_Unfiled", index=False)
        excluded.to_excel(writer, sheet_name="Excluded", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        criteria_frame().to_excel(writer, sheet_name="Criteria", index=False)
        criteria_text_frame.to_excel(writer, sheet_name="Full_Criteria_Text", index=False)
        examples_frame.to_excel(writer, sheet_name="Zotero_Examples", index=False)
        run_log.to_excel(writer, sheet_name="Run_Log", index=False)

    format_workbook(path)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    widths = {
        "Zotero Item Key": 16,
        "Title": 72,
        "AI Decision": 16,
        "AI Categorization": 38,
        "AI Confidence": 16,
        "AI Reasoning": 70,
        "Full Text Found?": 18,
        "Lexi Category Assignment": 28,
        "Lexi Notes": 42,
    }

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        headers = [cell.value for cell in sheet[1]]
        max_row = max(sheet.max_row, 2)
        if sheet_name in {"All_Unfiled", "Excluded"} and "Lexi Category Assignment" in headers:
            col = get_column_letter(headers.index("Lexi Category Assignment") + 1)
            validation = DataValidation(
                type="list",
                formula1=f"='Criteria'!$A$2:$A${len(allowed_categories()) + 1}",
                allow_blank=True,
            )
            sheet.add_data_validation(validation)
            validation.add(f"{col}2:{col}{max_row}")

        if sheet_name in {"All_Unfiled", "Excluded"}:
            for category_header in ("AI Categorization", "Lexi Category Assignment"):
                if category_header in headers:
                    col_index = headers.index(category_header) + 1
                    for row_num in range(2, max_row + 1):
                        value = str(sheet.cell(row=row_num, column=col_index).value or "")
                        category = value.split(";")[0].strip()
                        color = CATEGORY_COLORS.get(category, "FFFFFF")
                        sheet.cell(row=row_num, column=col_index).fill = PatternFill(
                            start_color=color,
                            end_color=color,
                            fill_type="solid",
                        )

        for idx, header in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = widths.get(
                str(header),
                42 if sheet_name in {"Criteria", "Full_Criteria_Text"} else 24,
            )
            if str(header) in {"Title", "AI Reasoning", "Lexi Notes", "Criteria Used", "Zotero Assignment Examples Used"}:
                for row_num in range(2, min(max_row, 250) + 1):
                    sheet.cell(row=row_num, column=idx).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate clean unfiled Zotero article category workbook.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Workbook output path.")
    parser.add_argument("--openai-model", default=os.getenv("OPENAI_CATEGORY_MODEL", DEFAULT_OPENAI_MODEL))
    parser.add_argument("--batch-size", type=int, default=8)
    parser.add_argument(
        "--criteria-file",
        type=Path,
        default=DEFAULT_CRITERIA_FILE if DEFAULT_CRITERIA_FILE.exists() else None,
        help="Text file containing updated inclusion/exclusion criteria.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    load_env_file()
    criteria_text = load_criteria_text(args.criteria_file)
    zotero_examples_text = fetch_zotero_assignment_examples()
    articles = fetch_unfiled_articles()
    articles = add_openai_categories(articles, args.openai_model, args.batch_size, criteria_text, zotero_examples_text)
    report = build_output_frame(articles)
    output_path = Path(args.output)
    write_workbook(output_path, report, args.openai_model, criteria_text, zotero_examples_text)
    print(f"Workbook written: {output_path.resolve()}")
    print(f"Rows: {len(report)}")


if __name__ == "__main__":
    main()
